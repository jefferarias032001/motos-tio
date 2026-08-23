#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
procesar_motos_tio.py
=====================
Lee "control_motos_tio.xlsx" (hoja "Registro Diario" + hoja "Clientes")
y genera el dashboard HTML.

Uso:
    python procesar_motos_tio.py

Con fecha manual (debug):
    python procesar_motos_tio.py --hoy 2026-07-28
"""

import argparse
import csv
import io
import json
import os
import sys
import urllib.request
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

NOMBRE_EXCEL   = "control_motos_tio.xlsx"
HOJA_REGISTRO  = "Registro Diario"
HOJA_CLIENTES  = "Clientes"
NOMBRE_NEGOCIO = "SEVERA MOTOS"    # ← cambia si quieres otro nombre en el dashboard
UMBRAL_ATRASO_LEVE = 2

# ---------------------------------------------------------------------------
# GOOGLE SHEETS — datos del cobrador en línea
# ---------------------------------------------------------------------------
# Cuando el cobrador llena el formulario (cobro_Code.gs), los pagos van a un
# Google Sheet.  Para que este script los lea automáticamente:
#   1. Abre el Google Sheet → Archivo → Compartir → Publicar en la web
#   2. Elige la hoja "Registro Diario" → formato CSV → Publicar
#   3. Pega el ID del Sheet en GOOGLE_SHEET_ID  (está en la URL del sheet)
#   4. Vuelve a correr: python procesar_motos_tio.py
# Si GOOGLE_SHEET_ID está vacío, el script solo usa el Excel local.
GOOGLE_SHEET_ID  = "2PACX-1vS0qMD_G2AcBCz4eErbdWfzllOwP0CloLc86X274_DGi4-XXszUEtQvylL9Xzr-z6AwvW4iUj_9o-Ag"
GOOGLE_SHEET_GID = "714608512"

# ---------------------------------------------------------------------------
# CATEGORÍAS DE CLIENTES
# clave parcial (substring del nombre en Excel) → configuración de pago
# cat: diario_30k | diario_34k | semanal | quincenal | mensual | sin_categoria
# dias: días del ciclo completo  |  meta: monto que completa una cuota
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# TELÉFONOS — agregar "tel": "3XXXXXXXXX" a cada cliente para habilitar SMS
# desde el dashboard cuando llevan 5+ días sin pagar.
# ---------------------------------------------------------------------------
CLIENTES_CONFIG = {
    "dorlys":           {"num":  1, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "elkin":            {"num":  2, "cat": "quincenal",   "dias": 15, "meta": 240_000, "label": "Quincenal 240k",  "tel": "3005106102"},
    "juan andres":      {"num":  3, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "jorge luis":       {"num":  4, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "henry junior":     {"num":  5, "cat": "diario_34k",  "dias": 7,  "meta": 238_000, "label": "Diario 34k × 7d"},
    "duvan enrique":    {"num":  6, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "wilmar ivis":      {"num":  7, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "jesus morales":    {"num":  8, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "asmed":            {"num":  9, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "chacal":           {"num": 10, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "wilian viejo":     {"num": 11, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "darwin":           {"num": 12, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "heiner rodriguez": {"num": 13, "cat": "diario_34k",  "dias": 7,  "meta": 240_000, "label": "Diario 34k × 7d"},
    "negro luis":       {"num": 14, "cat": "quincenal",   "dias": 15, "meta": 240_000, "label": "Quincenal 240k",   "tel": "3017986930"},
    "sr pedro":         {"num": 15, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "lucho laura":      {"num": 16, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "edinson":          {"num": 17, "cat": "diario_34k",  "dias": 7,  "meta": 238_000, "label": "Diario 34k × 7d",  "tel": "3017355501"},
    "brayan":           {"num": 18, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "gustavo primo":    {"num": 19, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "erik":             {"num": 20, "cat": "mensual",     "dias": 30, "meta": 420_000, "label": "Mensual 420k"},
    "efrain":           {"num": 21, "cat": "quincenal",   "dias": 15, "meta": 240_000, "label": "Quincenal 240k",  "tel": "3003111956"},
    "katherine":        {"num": 22, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "juan david":       {"num": 23, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "sr luis":          {"num": 24, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "manuel olga":      {"num": 25, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "jorge estrada":    {"num": 26, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "daniela":          {"num": 27, "cat": "mensual",     "dias": 30, "meta": 420_000, "label": "Mensual 420k",    "inicio": date(2026, 2, 6)},
    "alejandro":        {"num": 28, "cat": "diario_34k",  "dias": 7,  "meta": 238_000, "label": "Diario 34k × 7d"},
    "karluis":          {"num": 29, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "francisco":        {"num": 30, "cat": "quincenal",   "dias": 15, "meta": 190_000, "label": "Quincenal 190k"},
    "yesenia":          {"num": 31, "cat": "diario_34k",  "dias": 7,  "meta": 238_000, "label": "Diario 34k × 7d"},
    "alvania":          {"num": 32, "cat": "diario_34k",  "dias": 7,  "meta": 240_000, "label": "Diario 34k × 7d"},
    "yorkis":           {"num": 33, "cat": "semanal",     "dias": 7,  "meta": 240_000, "label": "Semanal 240k"},
    "josue":            {"num": 34, "cat": "quincenal",   "dias": 15, "meta": 260_000, "label": "Quincenal 260k"},
    "laura vanesa":     {"num": 35, "cat": "quincenal",   "dias": 15, "meta": 260_000, "label": "Quincenal 260k"},
    # Empresas externas (Luna / Jomar) — se procesan por separado pero necesitan config de ciclo
    "cesar":            {"num":  1, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
    "william villa":    {"num":  1, "cat": "diario_30k",  "dias": 8,  "meta": 240_000, "label": "Diario 30k × 8d"},
}

# Alias: nombre en app → nombre canónico en Excel (minúsculas)
ALIAS_NOMBRES = {
    "lucho / laura nosu":               "lucho laura nosu",
    "jesus morales / yuliana olivares":  "jesus morales yuliana olivares",
    "darwin / yaimis / ysabel":          "darwin yaimis ysabel",
    "ana milena-juan david":             "juan david-ana milena",
    "kevin rodriguez":                   "heiner rodriguez",
    "roberto":                           "edinson",
    "wilian junior / darlis esther":     "wilian viejo",
    "wilian junior":                     "wilian viejo",
    "chacal / caterine":                 "chacal caterine",
    # Negro: mismo cliente, diferentes nombres en app vs Excel
    "negro / maria / jose":              "negro luis ney",
    "negro maria jose":                  "negro luis ney",
}

# Nombre corto para mostrar en factura (clave → nombre mostrado)
NOMBRE_DISPLAY = {
    "wilian viejo":        "Wilian",
    "edinson el flaco":    "Edinson",
    "francisco (buñuelo)": "Francisco",
}

# Identificadores de empresas externas (substrings y placas en minúsculas)
# Usa coincidencia de substring para cubrir variantes de nombre en el Excel
_LUNA_SUBS   = ['cesar']
_LUNA_PLACAS = {'uyj15h'}
_JOMAR_SUBS  = ['william villa', 'william']   # 'william' ≠ 'wilian viejo' (doble l vs una l)
_JOMAR_PLACAS = {'bjr79i'}

def _empresa_de(clave, placa=None):
    """Retorna 'luna', 'jomar' o 'severa' según clave/placa del cliente."""
    p = str(placa).strip().lower() if placa else ''
    if any(s in clave for s in _LUNA_SUBS)  or p in _LUNA_PLACAS:  return 'luna'
    if any(s in clave for s in _JOMAR_SUBS) or p in _JOMAR_PLACAS: return 'jomar'
    return 'severa'

def normalizar_nombre(raw: str) -> str:
    """Normaliza variantes de nombres al nombre canónico."""
    key = raw.strip().lower()
    return ALIAS_NOMBRES.get(key, key)

def get_config(clave):
    """Busca la configuración de un cliente por substring de su clave."""
    c = clave.lower().strip()
    for k, cfg in CLIENTES_CONFIG.items():
        if k in c:
            return cfg
    return None

# ---------------------------------------------------------------------------
# MÓDULO LUNA — datos directos (sin Excel propio por ahora)
# ---------------------------------------------------------------------------
_LUNA_RAW = [
    {
        "nombre": "Cesar",
        "moto": "Moto Luna",
        "placa": "UYJ15H",
        "inicio": "2026-05-07",
        "total_cuotas": 64,
        "tarifa": 240_000,
        "pago_diario": 30_000,
        "observaciones": "Paga $30.000/dia x 8 dias = $240.000 por cuota",
        "pagos": [
            {"fecha": "2026-05-14", "valor": 240_000},
            {"fecha": "2026-05-22", "valor": 240_000},
            {"fecha": "2026-05-30", "valor": 240_000},
            {"fecha": "2026-06-07", "valor": 240_000},
            {"fecha": "2026-06-15", "valor": 240_000},
            {"fecha": "2026-06-23", "valor": 240_000},
            {"fecha": "2026-07-01", "valor": 240_000},
            {"fecha": "2026-07-09", "valor": 240_000},
            {"fecha": "2026-07-17", "valor": 240_000},
            {"fecha": "2026-07-25", "valor": 240_000},
            {"fecha": "2026-08-02", "valor": 240_000},
        ],
    }
]

# ---------------------------------------------------------------------------
# MÓDULO JOMAR — datos directos
# ---------------------------------------------------------------------------
_JOMAR_RAW = [
    {
        "nombre": "William Villa",
        "moto": "Motos Jomar",
        "placa": "BJR79I",
        "inicio": "2026-02-27",
        "total_cuotas": 72,
        "tarifa": 240_000,
        "pago_diario": 30_000,
        "telefono": "3207593591",
        "observaciones": "Paga $30.000/dia x 8 dias = $240.000 por cuota | Va puntual",
        "pagos": [
            {"fecha": "2026-03-06", "valor": 240_000},   # Cuota 1
            {"fecha": "2026-03-14", "valor": 240_000},   # Cuota 2
            {"fecha": "2026-03-22", "valor": 240_000},   # Cuota 3
            {"fecha": "2026-03-30", "valor": 240_000},   # Cuota 4
            {"fecha": "2026-04-07", "valor": 240_000},   # Cuota 5
            {"fecha": "2026-04-15", "valor": 240_000},   # Cuota 6
            {"fecha": "2026-04-23", "valor": 240_000},   # Cuota 7
            {"fecha": "2026-05-01", "valor": 240_000},   # Cuota 8
            {"fecha": "2026-05-09", "valor": 240_000},   # Cuota 9
            {"fecha": "2026-05-17", "valor": 240_000},   # Cuota 10
            {"fecha": "2026-05-25", "valor": 240_000},   # Cuota 11
            {"fecha": "2026-06-02", "valor": 240_000},   # Cuota 12
            {"fecha": "2026-06-10", "valor": 240_000},   # Cuota 13
            {"fecha": "2026-06-18", "valor": 240_000},   # Cuota 14
            {"fecha": "2026-06-26", "valor": 240_000},   # Cuota 15
            {"fecha": "2026-07-04", "valor": 240_000},   # Cuota 16
            {"fecha": "2026-07-12", "valor": 240_000},   # Cuota 17
            {"fecha": "2026-07-20", "valor": 240_000},   # Cuota 18
            {"fecha": "2026-07-28", "valor": 240_000},   # Cuota 19
            {"fecha": "2026-08-05", "valor": 240_000},   # Cuota 20 (hoy)
        ],
    }
]

def _construir_modulo(raw_list, hoy):
    clientes = []
    total_recibido = 0
    total_contratos = 0
    for c in raw_list:
        cuotas_pagadas   = len(c["pagos"])
        cuotas_restantes = c["total_cuotas"] - cuotas_pagadas
        recibido  = cuotas_pagadas  * c["tarifa"]
        pendiente = cuotas_restantes * c["tarifa"]
        total     = c["total_cuotas"] * c["tarifa"]
        ultimo_pago = c["pagos"][-1]["fecha"] if c["pagos"] else None
        dias_en_cuota = 0
        acumulado_cuota = 0
        if ultimo_pago:
            from datetime import date as _date
            ultimo = _date.fromisoformat(ultimo_pago)
            dias_en_cuota   = (hoy - ultimo).days
            acumulado_cuota = min(dias_en_cuota * c["pago_diario"], c["tarifa"])
        clientes.append({
            "nombre": c["nombre"], "moto": c["moto"], "placa": c["placa"],
            "inicio": c["inicio"], "total_cuotas": c["total_cuotas"],
            "cuotas_pagadas": cuotas_pagadas, "cuotas_restantes": cuotas_restantes,
            "tarifa": c["tarifa"], "pago_diario": c["pago_diario"],
            "recibido_historico": recibido, "pendiente_total": pendiente,
            "total_contrato": total, "ultimo_pago": ultimo_pago,
            "dias_en_cuota_actual": dias_en_cuota,
            "acumulado_cuota_actual": acumulado_cuota,
            "telefono": c.get("telefono", ""),
            "observaciones": c.get("observaciones", ""),
            "pagos": c["pagos"],
        })
        total_recibido  += recibido
        total_contratos += total
    return {
        "clientes": clientes,
        "global": {
            "recibido_historico": total_recibido,
            "pendiente_total": total_contratos - total_recibido,
            "total_contratos": total_contratos,
        },
    }

def construir_datos_luna(hoy):
    return _construir_modulo(_LUNA_RAW, hoy)

def construir_datos_jomar(hoy):
    return _construir_modulo(_JOMAR_RAW, hoy)

def _raw_a_filas(raw_list):
    """Convierte pagos hardcodeados (_LUNA_RAW/_JOMAR_RAW) a filas compatibles con procesar()."""
    filas = []
    for c in raw_list:
        clave = normalizar_nombre(c["nombre"])
        tarifa = c.get("tarifa", 240_000)
        for p in c.get("pagos", []):
            fecha = date.fromisoformat(p["fecha"]) if isinstance(p["fecha"], str) else p["fecha"]
            filas.append({
                "fecha":         fecha,
                "cliente_clave": clave,
                "cliente_raw":   c["nombre"],
                "moto":          c.get("moto"),
                "placa":         c.get("placa"),
                "pago_diario":   tarifa,
                "pago_recibido": float(p["valor"]),
                "medio_pago":    None,
                "saldo_cache":   None,
                "observaciones": "cobro_historico",
            })
    return filas

def _meta_de_raw(raw_list):
    """Construye un clientes_meta a partir de _LUNA_RAW/_JOMAR_RAW."""
    meta = {}
    for c in raw_list:
        clave = normalizar_nombre(c["nombre"])
        inicio_raw = c.get("inicio")
        if isinstance(inicio_raw, str):
            inicio_d = date.fromisoformat(inicio_raw)
        elif isinstance(inicio_raw, date):
            inicio_d = inicio_raw
        else:
            inicio_d = None
        meta[clave] = {
            "nombre":       c["nombre"],
            "moto":         c.get("moto"),
            "placa":        c.get("placa"),
            "inicio":       inicio_d,
            "observaciones": c.get("observaciones", ""),
            "total_cuotas": c.get("total_cuotas", 64),
            "telefono":     c.get("telefono", ""),
        }
    return meta

def _construir_datos_empresa(raw_list, filas_sheets, hoy):
    """Combina datos hardcodeados con cuotas nuevas de Google Sheets y procesa todo."""
    filas_hist = _raw_a_filas(raw_list)
    meta = _meta_de_raw(raw_list)
    if filas_sheets:
        ultimo_hist = max((f["fecha"] for f in filas_hist), default=date.min)
        filas_nuevas = [f for f in filas_sheets if f["fecha"] > ultimo_hist]
    else:
        filas_nuevas = []
    filas_total = filas_hist + filas_nuevas
    return procesar(filas_total, meta, hoy) if filas_total else None

MESES_ES = ["", "enero","febrero","marzo","abril","mayo","junio",
            "julio","agosto","septiembre","octubre","noviembre","diciembre"]
DIAS_ES       = ["lunes","martes","miércoles","jueves","viernes","sábado","domingo"]
DIAS_ES_ABREV = ["lun","mar","mié","jue","vie","sáb","dom"]

# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------

def formatear_pesos(valor):
    valor = round(valor or 0)
    signo = "-" if valor < 0 else ""
    return f"{signo}${abs(valor):,.0f}".replace(",",".")

def fecha_legible(d):
    return f"{DIAS_ES[d.weekday()].capitalize()} {d.day} de {MESES_ES[d.month]} de {d.year}"

def sumar_meses(d, n):
    """Suma n meses a una fecha manteniendo el día exacto (ajusta al último día del mes si hace falta)."""
    import calendar
    month = d.month + n
    year  = d.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    day   = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)

def resolver_ruta_excel():
    base = Path(__file__).resolve().parent
    ruta = base / NOMBRE_EXCEL
    if ruta.exists():
        return ruta
    print(f"No encontré '{NOMBRE_EXCEL}' junto al script ({base}).")
    print("Ejecuta 'python crear_datos_tio.py' primero para generarlo.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CARGA DE DATOS
# ---------------------------------------------------------------------------

def _tel_valido(raw):
    """Retorna el número si es móvil válido (colombiano 10 dígitos o internacional ≥11), sino ''."""
    s = str(raw).strip().replace(" ", "").replace("-", "").lstrip("+") if raw else ""
    if not s.isdigit():
        return ""
    if len(s) == 10 and s.startswith("3"):   # colombiano
        return s
    if len(s) >= 11:                          # internacional (ya trae código de país)
        return s
    return ""

def cargar_clientes(wb):
    """Retorna (info_por_nombre, tel_por_placa)."""
    info = {}
    tel_por_placa = {}
    if HOJA_CLIENTES not in wb.sheetnames:
        return info, tel_por_placa
    ws = wb[HOJA_CLIENTES]
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = (list(row) + [None]*7)[:7]
        nombre, moto, placa, inicio, obs, total_cuotas_raw, tel_raw = row
        if not nombre:
            continue
        clave = normalizar_nombre(str(nombre).strip().lower())
        tel = _tel_valido(tel_raw)
        info[clave] = {
            "nombre": str(nombre).strip().title(),
            "moto": moto,
            "placa": placa,
            "inicio": inicio.date() if isinstance(inicio, datetime) else None,
            "observaciones": obs,
            "total_cuotas": int(total_cuotas_raw) if isinstance(total_cuotas_raw, (int, float)) else 64,
            "telefono": tel,
        }
        if tel and placa:
            tel_por_placa[str(placa).strip().lower()] = tel
    return info, tel_por_placa

def _parsear_fecha(val):
    import calendar as _cal
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        s = val.strip().replace("//", "/")
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        # Fecha con día inválido (ej. 30/02/2026) → clampear al último día del mes
        for sep in ('/', '-'):
            if sep in s:
                parts = s.split(sep)
                if len(parts) == 3:
                    try:
                        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                        last = _cal.monthrange(y, m)[1]
                        return datetime(y, m, min(d, last))
                    except Exception:
                        pass
    return None

def cargar_registro(wb):
    ws = wb[HOJA_REGISTRO]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_raw = row[0] if len(row) > 0 else None
        cliente   = row[1] if len(row) > 1 else None
        fecha     = _parsear_fecha(fecha_raw)
        if not fecha or not cliente:
            continue
        pago_diario_raw = row[7] if len(row) > 7 else None
        pago_diario = float(pago_diario_raw) if isinstance(pago_diario_raw, (int,float)) else 0.0
        pago_recibido = row[8] if len(row) > 8 else None
        medio_pago    = row[9] if len(row) > 9 else None
        saldo_raw     = row[10] if len(row) > 10 else None
        observaciones = row[15] if len(row) > 15 else None
        cliente_raw = str(cliente).strip()
        filas.append({
            "fecha":          fecha.date(),
            "cliente_clave":  normalizar_nombre(cliente_raw),
            "cliente_raw":    cliente_raw,
            "moto":           row[2] if len(row) > 2 else None,
            "placa":          row[3] if len(row) > 3 else None,
            "pago_diario":    pago_diario,
            "pago_recibido":  float(pago_recibido) if isinstance(pago_recibido,(int,float)) else None,
            "medio_pago":     medio_pago if isinstance(medio_pago, str) else None,
            "saldo_cache":    float(saldo_raw) if isinstance(saldo_raw,(int,float)) else None,
            "observaciones":  str(observaciones).strip() if observaciones else None,
        })
    return filas

# ---------------------------------------------------------------------------
# GOOGLE SHEETS — lectura de pagos en línea
# ---------------------------------------------------------------------------

def cargar_registro_sheets(sheet_id, gid="0"):
    """Descarga el Registro Diario desde Google Sheets (hoja publicada como CSV)."""
    url = (
        f"https://docs.google.com/spreadsheets/d/e/{sheet_id}"
        f"/pub?output=csv&gid={gid}"
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            text = resp.read().decode("utf-8-sig")
    except Exception as e:
        print(f"⚠️  No se pudo leer Google Sheets: {e}")
        return []

    filas = []
    # DictReader usa la primera fila como encabezado → resistente a cambios de columnas
    reader = csv.DictReader(io.StringIO(text))

    def _num(v):
        try:
            return float(str(v).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            return None

    def _get(row, *keys):
        """Busca la primera clave que exista y tenga valor."""
        for k in keys:
            v = row.get(k, "").strip()
            if v:
                return v
        return ""

    for row in reader:
        fecha_raw = _get(row, "fecha", "Fecha", "FECHA")
        cliente   = _get(row, "cliente", "Cliente", "CLIENTE")
        if not fecha_raw or not cliente:
            continue
        # Parsear fecha (Google Sheets exporta como M/D/YYYY o YYYY-MM-DD o D/M/YYYY)
        fecha = None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                fecha = datetime.strptime(fecha_raw.split()[0], fmt.split()[0]).date()
                break
            except ValueError:
                continue
        if fecha is None:
            continue

        pago_diario_raw   = _get(row, "pago_diario",   "Pago Diario",   "PAGO DIARIO")
        pago_recibido_raw = _get(row, "pago_recibido", "Pago Recibido", "PAGO RECIBIDO")
        medio_pago        = _get(row, "medio_pago",    "Medio Pago",    "MEDIO PAGO") or None
        observaciones     = _get(row, "observaciones", "Observaciones", "OBSERVACIONES") or None
        moto              = _get(row, "moto",  "Moto",  "MOTO")  or None
        placa             = _get(row, "placa", "Placa", "PLACA") or None

        pago_diario   = _num(pago_diario_raw)
        pago_recibido = _num(pago_recibido_raw)

        cliente_raw = cliente.strip()
        filas.append({
            "fecha":          fecha,
            "cliente_clave":  normalizar_nombre(cliente_raw),
            "cliente_raw":    cliente_raw,
            "moto":           moto,
            "placa":          placa,
            "pago_diario":    pago_diario or 0.0,
            "pago_recibido":  pago_recibido,
            "medio_pago":     medio_pago,
            "saldo_cache":    None,
            "observaciones":  observaciones,
        })

    print(f"📡 Google Sheets: {len(filas)} pagos cargados online.")
    return filas


def agrupar_en_cuotas(filas_online, tarifa_default=240_000):
    """
    Agrupa pagos diarios de Google Sheets en cuotas completas.
    Cada vez que el acumulado de un cliente llega a su tarifa ($240.000),
    se genera UNA sola entrada de cuota en lugar de N entradas diarias.
    Las cuotas incompletas (ciclo en curso) no se agregan.
    """
    por_cliente = defaultdict(list)
    for f in sorted(filas_online, key=lambda r: r["fecha"]):
        por_cliente[f["cliente_clave"]].append(f)

    cuotas = []
    for clave, pagos in por_cliente.items():
        # Prioridad 1: usar meta de CLIENTES_CONFIG (es la fuente correcta de verdad)
        cfg = get_config(clave)
        tarifa = cfg["meta"] if cfg else tarifa_default

        # Prioridad 2: inferir desde pago_diario solo si no hay config
        if not cfg:
            tarifas = [p["pago_diario"] for p in pagos if p["pago_diario"] and p["pago_diario"] > 0]
            if tarifas:
                diario = max(set(tarifas), key=tarifas.count)
                tarifa = diario * 8 if diario <= 50_000 else diario

        acum = 0.0
        plantilla = None  # primer pago del ciclo (para copiar metadatos)
        for p in pagos:
            recibido = p["pago_recibido"] or 0.0
            if recibido <= 0:
                continue
            if plantilla is None:
                plantilla = p
            acum += recibido
            while acum >= tarifa:
                excedente = acum - tarifa
                cuota_entry = {
                    **plantilla,
                    "fecha":          p["fecha"],
                    "pago_diario":    tarifa,
                    "pago_recibido":  tarifa,
                    # solo mostrar excedente si no alcanza para otra cuota (no genera chip "sdo" falso)
                    "excedente":      excedente if excedente < tarifa else 0,
                    "saldo_cache":    None,
                    "observaciones":  "cobro_app",
                }
                cuotas.append(cuota_entry)
                acum = max(0.0, excedente)
                plantilla = p if acum > 0 else None

    print(f"   → {len(cuotas)} cuotas completas agrupadas desde Google Sheets")
    return cuotas

# ---------------------------------------------------------------------------
# PROCESAMIENTO  (idéntico a la versión de Jeffer)
# ---------------------------------------------------------------------------

def saldo_en_o_antes(historial_saldo, fecha_corte):
    vigente = 0.0
    for f, s in historial_saldo:
        if f > fecha_corte:
            break
        vigente = s
    return vigente

def calcular_estado(saldo, tarifa_ref):
    if saldo <= 0:
        return "al_dia"
    tarifas = saldo / tarifa_ref if tarifa_ref else 0
    return "atraso_leve" if tarifas <= UMBRAL_ATRASO_LEVE else "atraso_alto"

def frase_resumen_semana(nombre, dias_atrasados, monto):
    if not dias_atrasados:
        return f"Esta semana {nombre} está al día."
    nombres_dias = [DIAS_ES[d.weekday()] for d in sorted(dias_atrasados)]
    texto_dias = (", ".join(nombres_dias[:-1]) + " y " + nombres_dias[-1]) if len(nombres_dias) > 1 else nombres_dias[0]
    plural = "día" if len(nombres_dias) == 1 else "días"
    return f"Esta semana {nombre} se quedó {len(nombres_dias)} {plural} ({texto_dias}) por {formatear_pesos(monto)}."

def procesar(filas, clientes_meta, hoy, tel_por_placa=None, ultimos_sheets=None):
    por_cliente = defaultdict(list)
    for f in filas:
        por_cliente[f["cliente_clave"]].append(f)

    resultado_clientes = []
    tendencia = defaultdict(lambda: {"esperado": 0.0, "recibido": 0.0})
    tabla_detalle = []
    total_global = defaultdict(float)
    ultimo_dia_mes_anterior = hoy.replace(day=1) - timedelta(days=1)

    for clave, registros in por_cliente.items():
        registros.sort(key=lambda r: r["fecha"])
        meta = clientes_meta.get(clave, {})
        cfg  = get_config(clave)   # configuración de categoría/ciclo
        nombre = registros[0]["cliente_raw"].title() or meta.get("nombre", "")
        moto   = meta.get("moto")   or registros[0]["moto"]
        placa  = meta.get("placa")  or registros[0]["placa"]
        inicio = (cfg.get("inicio") if cfg else None) or meta.get("inicio") or registros[0]["fecha"]

        tarifas_validas = [r["pago_diario"] for r in registros if r["pago_diario"] > 0]
        tarifa_ref = Counter(tarifas_validas).most_common(1)[0][0] if tarifas_validas else 240_000.0

        saldo_efectivo = 0.0
        historial_saldo = []
        historial = {}
        recibido_total = 0.0
        recibido_mes   = 0.0
        ultimo_pago_fecha = None

        for r in registros:
            if r["fecha"] > hoy:
                continue
            recibido = r["pago_recibido"] or 0.0
            if r["saldo_cache"] is not None:
                saldo_efectivo = r["saldo_cache"]
            else:
                saldo_efectivo += r["pago_diario"] - recibido
            historial_saldo.append((r["fecha"], saldo_efectivo))
            historial[r["fecha"]] = r
            recibido_total += recibido
            if r["fecha"].year == hoy.year and r["fecha"].month == hoy.month:
                recibido_mes += recibido
            if recibido > 0:
                ultimo_pago_fecha = r["fecha"]
            # No incluir cuotas agrupadas de la app — el detalle muestra pagos reales
            if r.get("observaciones") == "cobro_app":
                continue
            tabla_detalle.append({
                "fecha":      r["fecha"].isoformat(),
                "cliente":    nombre,
                "placa":      placa,
                "esperado":   r["pago_diario"],
                "recibido":   recibido,
                "diferencia": r["pago_diario"] - recibido,
                "saldo":      saldo_efectivo,
                "medio_pago": r["medio_pago"],
                "nota":       r["observaciones"],
            })

        saldo_atrasado = historial_saldo[-1][1] if historial_saldo else 0.0
        fecha_fin_contrato = max((r["fecha"] for r in registros), default=hoy)
        restante_programado = sum(max(0.0, r["pago_diario"]) for r in registros if r["fecha"] > hoy)
        deuda_total_con_programado = saldo_atrasado + restante_programado
        atrasado_mes = saldo_atrasado - saldo_en_o_antes(historial_saldo, ultimo_dia_mes_anterior)

        esperado_total = recibido_total + saldo_atrasado
        esperado_mes   = recibido_mes + atrasado_mes

        # --- métricas semanales / contrato ---
        # Overrides manuales de total_cuotas (cuando el Excel trae un valor incorrecto)
        CUOTAS_OVERRIDE = {
            "dorlys jose julio": 60,
            "henry junior":      52,
            "josue alzazar":     48,
            "laura vanesa":      48,
        }
        total_cuotas = CUOTAS_OVERRIDE.get(clave, meta.get("total_cuotas", 64))
        pagos_historicos = sorted([r for r in registros if r["fecha"] <= hoy], key=lambda r: r["fecha"])
        pagos_realizados = len(pagos_historicos)
        # Si hubo un pago en bloque que cubre varias cuotas, contar por monto
        meta_cuota = cfg["meta"] if cfg else 240_000
        if meta_cuota > 0:
            cuotas_por_monto = int(recibido_total // meta_cuota)
            if cuotas_por_monto > pagos_realizados:
                pagos_realizados = cuotas_por_monto
        pagos_restantes  = max(0, total_cuotas - pagos_realizados)
        porc_completado  = round(pagos_realizados / total_cuotas * 100, 1) if total_cuotas > 0 else 0

        # fecha fin de contrato = inicio contrato + total_cuotas ciclos
        dias_ciclo_contrato = cfg["dias"] if cfg else 7
        cat_ciclo = cfg["cat"] if cfg else ""
        if cat_ciclo in ("mensual", "quincenal"):
            # sumar meses/quincenas exactas en lugar de días fijos
            meses_por_ciclo = 1 if cat_ciclo == "mensual" else 0
            quincenas = total_cuotas if cat_ciclo == "quincenal" else 0
            if cat_ciclo == "mensual":
                fecha_fin_prevista = sumar_meses(inicio, total_cuotas)
            else:  # quincenal ≈ 15 días exactos
                fecha_fin_prevista = inicio + timedelta(days=total_cuotas * 15)
        else:
            fecha_fin_prevista = inicio + timedelta(days=total_cuotas * dias_ciclo_contrato)

        # contrato completado?
        contrato_completo = pagos_realizados >= total_cuotas

        # días sin pagar — combinar último pago Excel + último pago raw Sheets
        # (los pagos parciales del ciclo en curso no llegan como cuota completa)
        dias_sin_pagar = 0
        proximo_pago_esperado = None
        dias_ciclo_cfg = cfg["dias"] if cfg else 7
        cat_ciclo_str  = cfg["cat"] if cfg else ""
        ultimo_raw_sheets = (ultimos_sheets or {}).get(clave)
        ultimo_pago_real = max(
            [d for d in [ultimo_pago_fecha, ultimo_raw_sheets] if d],
            default=None
        )
        if not contrato_completo and ultimo_pago_real:
            proximo_pago_esperado = ultimo_pago_real + timedelta(days=dias_ciclo_cfg)
            # Diario: contar días sueltos sin pago (ciclo=1), no el ciclo de la cuota
            if cat_ciclo_str.startswith("diario"):
                prox_sinpagar = ultimo_pago_real + timedelta(days=1)
            else:
                prox_sinpagar = proximo_pago_esperado
            if prox_sinpagar < hoy:
                dias_sin_pagar = (hoy - prox_sinpagar).days

        # últimos 8 pagos semanales para mostrar en la tarjeta
        ultimos_8 = pagos_historicos[-8:]
        base_num  = pagos_realizados - len(ultimos_8) + 1
        ultimos_pagos_data = []
        for i_p, r in enumerate(ultimos_8):
            rec = r["pago_recibido"] or 0
            esp = r["pago_diario"]
            if rec >= esp and esp > 0:
                est = "pagado"
            elif rec > 0:
                est = "parcial"
            else:
                est = "no_pago"
            # Para entradas de cobro_app: mostrar la meta exacta, no el total acumulado
            excedente = r.get("excedente", 0) or 0
            rec_display = esp if (r.get("observaciones") == "cobro_app" and excedente > 0) else rec
            ultimos_pagos_data.append({
                "fecha":     r["fecha"].isoformat(),
                "mes_dia":   f"{r['fecha'].day}/{r['fecha'].month}",
                "dia_nombre": DIAS_ES_ABREV[r["fecha"].weekday()],
                "recibido":  rec_display,
                "esperado":  esp,
                "estado":    est,
                "num_cuota": base_num + i_p,
                "excedente": round(excedente),
            })

        meses_con_actividad = sorted({(r["fecha"].year, r["fecha"].month) for r in registros if r["fecha"] <= hoy})
        for anio_m, mes_m in meses_con_actividad:
            inicio_mes = date(anio_m, mes_m, 1)
            fin_mes = date(anio_m, mes_m+1, 1) - timedelta(days=1) if mes_m < 12 else date(anio_m+1, 1, 1) - timedelta(days=1)
            corte_fin = min(fin_mes, hoy)
            recibido_del_mes = sum(
                (r["pago_recibido"] or 0.0)
                for r in registros
                if r["fecha"] <= hoy and r["fecha"].year == anio_m and r["fecha"].month == mes_m
            )
            saldo_fin_mes = saldo_en_o_antes(historial_saldo, corte_fin)
            saldo_fin_mes_anterior = saldo_en_o_antes(historial_saldo, inicio_mes - timedelta(days=1))
            clave_mes = (anio_m, mes_m)
            tendencia[clave_mes]["recibido"] += recibido_del_mes
            tendencia[clave_mes]["esperado"] += recibido_del_mes + (saldo_fin_mes - saldo_fin_mes_anterior)

        semana = []
        dias_atrasados_semana = []
        monto_atrasado_semana = 0.0
        for i in range(6, -1, -1):
            d = hoy - timedelta(days=i)
            if d < inicio:
                continue
            r = historial.get(d)
            if r is None:
                estado_dia = "sin_dato"
                esperado_d = recibido_d = nota_d = None
            else:
                esperado_d = r["pago_diario"]
                recibido_d = r["pago_recibido"]
                nota_d     = r["observaciones"]
                if esperado_d == 0 and recibido_d:
                    estado_dia = "pagado"
                elif esperado_d == 0:
                    estado_dia = "sin_cobro"
                elif d == hoy and recibido_d is None:
                    estado_dia = "pendiente"
                elif recibido_d is None or recibido_d == 0:
                    estado_dia = "no_pago"
                elif recibido_d < esperado_d:
                    estado_dia = "parcial"
                else:
                    estado_dia = "pagado"
                if estado_dia in ("no_pago", "parcial"):
                    dias_atrasados_semana.append(d)
                    monto_atrasado_semana += esperado_d - (recibido_d or 0)
            semana.append({
                "fecha": d.isoformat(), "dia_nombre": DIAS_ES_ABREV[d.weekday()],
                "dia_numero": d.day, "esperado": esperado_d, "recibido": recibido_d,
                "estado": estado_dia, "nota": nota_d, "es_hoy": d == hoy,
            })

        cumplimiento  = (recibido_total / esperado_total * 100) if esperado_total > 0 else 100.0
        if contrato_completo:
            estado_general = "finalizado"
        elif dias_sin_pagar > 14:
            estado_general = "atraso_alto"
        elif dias_sin_pagar > 0:
            estado_general = "atraso_leve"
        else:
            estado_general = calcular_estado(saldo_atrasado, tarifa_ref)

        resultado_clientes.append({
            "clave": clave, "nombre": nombre, "moto": moto, "placa": placa,
            "inicio": inicio.isoformat(), "inicio_legible": fecha_legible(inicio),
            "tarifa_ref": tarifa_ref,
            "saldo_atrasado": saldo_atrasado, "saldo_en_tarifas": round(saldo_atrasado/tarifa_ref,1) if tarifa_ref else 0,
            "atrasado_mes": atrasado_mes, "recibido_total": recibido_total,
            "recibido_mes": recibido_mes, "esperado_total": esperado_total,
            "esperado_mes": esperado_mes, "restante_programado": restante_programado,
            "fecha_fin_contrato": fecha_fin_contrato.isoformat(),
            "fecha_fin_contrato_legible": fecha_legible(fecha_fin_contrato),
            "deuda_total_con_programado": deuda_total_con_programado,
            "cumplimiento": round(cumplimiento, 1),
            "ultimo_pago_fecha": ultimo_pago_fecha.isoformat() if ultimo_pago_fecha else None,
            "estado_general": estado_general, "semana": semana,
            "resumen_semana": frase_resumen_semana(nombre, dias_atrasados_semana, monto_atrasado_semana),
            "dias_atrasados_semana": len(dias_atrasados_semana),
            "monto_atrasado_semana": monto_atrasado_semana,
            # campos semanales / contrato
            "total_cuotas":    total_cuotas,
            "pagos_realizados": pagos_realizados,
            "pagos_restantes":  pagos_restantes,
            "porc_completado":  porc_completado,
            "fecha_fin_prevista": fecha_fin_prevista.isoformat(),
            "fecha_fin_prevista_legible": fecha_legible(fecha_fin_prevista),
            "contrato_completo": contrato_completo,
            "dias_sin_pagar":   dias_sin_pagar,
            "ultimo_pago":      ultimo_pago_real.isoformat() if ultimo_pago_real else None,
            "proximo_pago_esperado": proximo_pago_esperado.isoformat() if proximo_pago_esperado else None,
            "proximo_pago_esperado_legible": fecha_legible(proximo_pago_esperado) if proximo_pago_esperado else "—",
            "ultimos_pagos":    ultimos_pagos_data,
            # ── Categoría y proyección ──────────────────────────────────────
            "numero":           cfg["num"]   if cfg else 999,
            "categoria":        cfg["cat"]   if cfg else "sin_categoria",
            "tipo_pago_label":  cfg["label"] if cfg else "FALTAN POR CATEGORIA",
            "dias_ciclo":       cfg["dias"]  if cfg else 7,
            "meta_ciclo":       cfg["meta"]  if cfg else 240_000,
            "prox_cuota_fecha": proximo_pago_esperado.isoformat() if proximo_pago_esperado else None,
            "prox_cuota_dias":  (proximo_pago_esperado - hoy).days if proximo_pago_esperado else None,
            "telefono": (tel_por_placa or {}).get(str(placa).strip().lower(), "") or meta.get("telefono", "") or (cfg.get("tel", "") if cfg else ""),
        })

        total_global["recibido_historico"]  += recibido_total
        total_global["recibido_mes"]        += recibido_mes
        total_global["esperado_historico"]  += esperado_total
        total_global["esperado_mes"]        += esperado_mes
        total_global["atrasado_historico"]  += saldo_atrasado
        total_global["atrasado_mes"]        += atrasado_mes
        total_global["restante_programado_total"] += restante_programado

    # Orden fijo por número de categoría (finalizados al final)
    resultado_clientes.sort(key=lambda c: (1 if c["contrato_completo"] else 0, c["numero"]))

    dias_en_mes_actual = (date(hoy.year, hoy.month+1, 1) - timedelta(days=1)).day if hoy.month < 12 else 31
    total_global["cantidad_motos"] = len({c["moto"] for c in resultado_clientes if c["moto"]}) or len(resultado_clientes)
    total_global["tarifa_diaria_combinada"] = sum(c["tarifa_ref"] for c in resultado_clientes)
    total_global["esperado_mes_completo"]   = total_global["tarifa_diaria_combinada"] * dias_en_mes_actual
    total_global["deuda_total_con_programado"] = total_global["atrasado_historico"] + total_global["restante_programado_total"]
    total_global["cumplimiento"] = round(
        (total_global["recibido_historico"] / total_global["esperado_historico"] * 100)
        if total_global["esperado_historico"] > 0 else 100.0, 1)

    tendencia_lista = [
        {"anio": a, "mes": m, "mes_nombre": MESES_ES[m][:3], "esperado": v["esperado"], "recibido": v["recibido"]}
        for (a, m), v in sorted(tendencia.items())
    ]
    tabla_detalle.sort(key=lambda x: x["fecha"], reverse=True)

    return {"clientes": resultado_clientes, "global": dict(total_global),
            "tendencia": tendencia_lista, "detalle": tabla_detalle}

# ---------------------------------------------------------------------------
# SALIDA
# ---------------------------------------------------------------------------

def imprimir_resumen_terminal(datos, hoy):
    banderas = {"al_dia": "[OK]", "atraso_leve": "[!] ", "atraso_alto": "[!!]", "finalizado": "[✓] "}
    print()
    print("=" * 64)
    print(f" {NOMBRE_NEGOCIO} · resumen al {hoy.strftime('%d/%m/%Y')}")
    print("=" * 64)
    for c in datos["clientes"]:
        b = banderas[c["estado_general"]]
        print(f" {b} {c['nombre']:<22} atrasado: {formatear_pesos(c['saldo_atrasado']):>13}")
        if c["dias_atrasados_semana"]:
            print(f"        -> {c['resumen_semana']}")
    print("-" * 64)
    g = datos["global"]
    print(f" Recibido histórico : {formatear_pesos(g['recibido_historico'])}")
    print(f" Atrasado total     : {formatear_pesos(g['atrasado_historico'])}")
    print(f" Cumplimiento       : {g['cumplimiento']}%")
    print("=" * 64)

def generar_html(datos, ruta_plantilla, ruta_salida, hoy,
                 datos_luna_excel=None, datos_jomar_excel=None):
    plantilla = ruta_plantilla.read_text(encoding="utf-8")
    ts = datetime.now().strftime("%d/%m/%Y %I:%M %p")
    payload = {
        "generado_en": ts,
        "hoy": hoy.isoformat(), "hoy_legible": fecha_legible(hoy),
        **datos,
    }
    # Luna: usa datos del Excel si están disponibles, sino los hardcodeados
    if datos_luna_excel:
        datos_luna = {"generado_en": ts, "hoy": hoy.isoformat(),
                      "hoy_legible": fecha_legible(hoy), **datos_luna_excel}
    else:
        datos_luna = construir_datos_luna(hoy)
    # Jomar: igual
    if datos_jomar_excel:
        datos_jomar = {"generado_en": ts, "hoy": hoy.isoformat(),
                       "hoy_legible": fecha_legible(hoy), **datos_jomar_excel}
    else:
        datos_jomar = construir_datos_jomar(hoy)
    # Construir dict WhatsApp desde todos los clientes (Severa + Luna + Jomar)
    wa_dict = {}
    for c in payload.get("clientes", []):
        tel = str(c.get("telefono", "") or "").replace(" ", "").replace("-", "")
        tel = ''.join(filter(str.isdigit, tel))
        if tel and len(tel) >= 7:
            key = str(c.get("nombre", "")).lower().strip()
            wa_dict[key] = tel if len(tel) >= 11 else f"57{tel}"
    for modulo in [datos_luna, datos_jomar]:
        for c in modulo.get("clientes", []):
            tel = str(c.get("telefono", "") or "").replace(" ", "").replace("-", "")
            tel = ''.join(filter(str.isdigit, tel))
            if tel and len(tel) >= 7:
                key = str(c.get("nombre", "")).lower().strip()
                wa_dict[key] = tel if len(tel) >= 11 else f"57{tel}"

    html = plantilla.replace("__DATOS_JSON_AQUI__", json.dumps(payload, ensure_ascii=False))
    html = html.replace("__DATOS_LUNA_JSON__",  json.dumps(datos_luna,  ensure_ascii=False))
    html = html.replace("__DATOS_JOMAR_JSON__", json.dumps(datos_jomar, ensure_ascii=False))
    html = html.replace("__WHATSAPP_CLIENTES_JSON__", json.dumps(wa_dict, ensure_ascii=False))
    html = html.replace("__NOMBRE_DISPLAY_JSON__", json.dumps(NOMBRE_DISPLAY, ensure_ascii=False))
    html = html.replace("JEFFER MOTOS", NOMBRE_NEGOCIO)
    ruta_salida.write_text(html, encoding="utf-8")
    # También guardar como index.html para que GitHub Pages lo sirva en la raíz
    ruta_index = ruta_salida.parent / "index.html"
    ruta_index.write_text(html, encoding="utf-8")

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--hoy", help="AAAA-MM-DD (debug)")
    parser.add_argument("--salida", help="Ruta del HTML de salida")
    args = parser.parse_args()

    ruta_excel = resolver_ruta_excel()
    hoy = datetime.strptime(args.hoy, "%Y-%m-%d").date() if args.hoy else date.today()

    print(f"Leyendo: {ruta_excel}")
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    clientes_meta, tel_por_placa = cargar_clientes(wb)
    filas = cargar_registro(wb)
    print(f"{len(filas)} filas del Excel ({len(clientes_meta)} clientes, {len(tel_por_placa)} con tel)")

    # Combinar con Google Sheets si está configurado
    # Último pago real por cliente en raw Sheets (sin agrupar) — para calcular "al día" correctamente
    ultimos_sheets = {}
    if GOOGLE_SHEET_ID:
        filas_online = cargar_registro_sheets(GOOGLE_SHEET_ID, GOOGLE_SHEET_GID)
        if filas_online:
            # Capturar último pago raw antes de agrupar
            for f in filas_online:
                if (f.get("pago_recibido") or 0) > 0:
                    clave = f["cliente_clave"]
                    if clave not in ultimos_sheets or f["fecha"] > ultimos_sheets[clave]:
                        ultimos_sheets[clave] = f["fecha"]
            # Agrupar pagos diarios en cuotas completas antes de mezclar
            cuotas_online = agrupar_en_cuotas(filas_online)
            # Solo agregar cuotas de app POSTERIORES al último pago Excel del cliente
            # Evita doble conteo cuando el cobrador ya registró en Excel lo que la app también tiene
            ultimo_excel = {}
            for f in filas:
                clave = f["cliente_clave"]
                if clave not in ultimo_excel or f["fecha"] > ultimo_excel[clave]:
                    ultimo_excel[clave] = f["fecha"]
            nuevos = [f for f in cuotas_online
                      if f["fecha"] > ultimo_excel.get(f["cliente_clave"], date.min)]
            filas.extend(nuevos)
            print(f"   + {len(nuevos)} cuotas nuevas desde Google Sheets")
    print(f"Total: {len(filas)} registros combinados")

    # Separar filas y clientes por empresa
    filas_severa = [f for f in filas if _empresa_de(f["cliente_clave"], f.get("placa")) == 'severa']
    filas_luna   = [f for f in filas if _empresa_de(f["cliente_clave"], f.get("placa")) == 'luna']
    filas_jomar  = [f for f in filas if _empresa_de(f["cliente_clave"], f.get("placa")) == 'jomar']
    clientes_severa = {k: v for k, v in clientes_meta.items() if _empresa_de(k, v.get("placa")) == 'severa'}
    clientes_luna   = {k: v for k, v in clientes_meta.items() if _empresa_de(k, v.get("placa")) == 'luna'}
    clientes_jomar  = {k: v for k, v in clientes_meta.items() if _empresa_de(k, v.get("placa")) == 'jomar'}
    print(f"  Severa: {len(filas_severa)} reg / {len(clientes_severa)} clientes")
    print(f"  Luna:   {len(filas_luna)} reg / {len(clientes_luna)} clientes")
    print(f"  Jomar:  {len(filas_jomar)} reg / {len(clientes_jomar)} clientes")

    datos = procesar(filas_severa, clientes_severa, hoy, tel_por_placa, ultimos_sheets)
    datos_luna_exc  = _construir_datos_empresa(_LUNA_RAW,  filas_luna,  hoy)
    datos_jomar_exc = _construir_datos_empresa(_JOMAR_RAW, filas_jomar, hoy)

    base_dir = Path(__file__).resolve().parent
    # Usa la plantilla tío (semanal); si no existe, cae al original
    ruta_plantilla = base_dir / "plantilla_dashboard_tio.html"
    if not ruta_plantilla.exists():
        ruta_plantilla = base_dir.parent / "control-motos" / "plantilla_dashboard.html"
    if not ruta_plantilla.exists():
        print("No encontré plantilla_dashboard.html.")
        sys.exit(1)

    ruta_salida = Path(args.salida) if args.salida else base_dir / "dashboard_motos_tio.html"
    generar_html(datos, ruta_plantilla, ruta_salida, hoy,
                 datos_luna_excel=datos_luna_exc, datos_jomar_excel=datos_jomar_exc)
    imprimir_resumen_terminal(datos, hoy)
    print(f"\nDashboard generado en: {ruta_salida}")

if __name__ == "__main__":
    main()
