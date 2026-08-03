#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
crear_datos_tio.py
==================
Genera el archivo Excel con los pagos de las dos motos
extraídos de las fotos del cuaderno.

Ejecución:
    python crear_datos_tio.py
"""

import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Falta openpyxl. Instálalo con:  pip install openpyxl")
    sys.exit(1)

SALIDA = Path(__file__).resolve().parent / "control_motos_tio.xlsx"

# ---------------------------------------------------------------------------
# DATOS EXTRAÍDOS DE LAS FOTOS
# Nota: algunas fechas pueden tener ±1-2 días de imprecisión por la letra
#       manuscrita. El usuario puede corregirlas directamente en el Excel.
# ---------------------------------------------------------------------------

TARIFA = 240_000  # cuota semanal fija (pesos)

CLIENTES = [
    # nombre, moto, placa, fecha_inicio, observaciones, total_cuotas
    ("Dorlys Jose Julio",     "Bajaj EYO 114", "EYO114", datetime(2024,12,18), "Tel: 3205376845 | 595000 inicial", 64),
    ("Elkin Espitia Payares", "Bajaj EYO 06H", "EYO06H", datetime(2024,12,26), "Tel: 3003xx | 5980000",           64),
    ("Juan Andres De Arco",   "Bajaj KOO 16H", "KOO16H", datetime(2025, 1, 5), "Tel: 3202386294 | 5950000 inicial | 63 pagos (1 pendiente)", 64),
    ("Jorge Luis Reyes",      "Bajaj KOO B1H", "KOOB1H", datetime(2025, 1,10), "Tel: 3202841723 / 3005685892 | 5950000 inicial | CONTRATO FINALIZADO", 64),
    ("Luis Gabriel Polo",     "Bajaj KOM 89H", "KOM89H", datetime(2025,12,23), "Tel: 3005685892 | 6806000 inicial | 72 cuotas/18 meses | Moto anterior OYL39H robada, repuesta dic-2025", 72),
    ("Duvon Enrique",         "Bajaj KEB 58H", "KEB58H", datetime(2025, 2,17), "Tel: 3126959090 | 15360000 total | 16 meses | 63 pagos (1 pendiente)", 64),
    ("Wilmer Ivis",           "Bajaj KEW 23H", "KEW23H", datetime(2025, 4, 4), "Tel: 3233311964 | 15360000 total | 64 cuotas | 59 pagos (5 pendientes)", 64),
    ("Jesus Morales Yuliana Olivares", "Bajaj KFK 1AH", "KFK1AH", datetime(2025, 5, 8), "Tel: 3465625710 | 17280000 total | 18 meses | 52 pagos (20 pendientes)", 72),
    ("Osmed Brenda",                  "Bajaj OSZ 46H", "OSZ46H", datetime(2025, 5,15), "Tel: 3205197351 | 15360000 total | 64 cuotas | 54 pagos (10 pendientes) | ATRASADO 21 dias", 64),
    ("Chacal Caterine",               "Bajaj OSZ 43H", "OSZ43H", datetime(2025, 5,13), "Tel: 3245780119 | 17280000 total | 72 cuotas | 50 pagos (22 pendientes)", 72),
    ("Wilian Junior Darlis Esther",   "Bajaj OTE 02H", "OTE02H", datetime(2025, 5,22), "Tel: 3239896985 | 14400000 total | 60 cuotas | 34 pagos (26 pendientes)", 60),
    ("Darwin Yaimis Ysabel",          "Bajaj OTE 80H", "OTE80H", datetime(2025, 5,25), "15360000 total | 64 cuotas | 54 pagos (10 pendientes)", 64),
    ("Kevin Rodriguez",               "Bajaj OXO 04H", "OXO04H", datetime(2025, 5,27), "20160000 total | 84 cuotas/21 meses | pagos bisemanales | 29 pagos (55 pendientes)", 84),
    ("Nabro Maria Jesse",             "Bajaj OTE 68H", "OTE68H", datetime(2025, 4,22), "15360000 total | 64 cuotas | 51 pagos (13 pendientes)", 64),
    ("Luis Katherin Estor",           "Bajaj OTG 82H", "OTG82H", datetime(2025, 6, 4), "Fotos 27-28 | Ref: Katherin Estor 3137781909 | 69 cuotas | 54 pagos (15 pendientes)", 69),
    ("Lucho Laura Nosu",             "Bajaj OTF 59H", "OTF59H", datetime(2025, 5,31), "Tel: 3147356412 | 15360000 total | 64 cuotas | 51 pagos (13 pendientes)", 64),
    ("Roberto",                      "Bajaj OTR 29H", "OTR29H", datetime(2025, 6,24), "Ref: Nik amigo De Guillermo / Flaco Hobiter Edinson | 15360000 total | 64 cuotas | 43 pagos (21 pendientes)", 64),
    ("Braillon",                     "Bajaj OUA 19H", "OUA19H", datetime(2025, 7,17), "15360000 total | 64 cuotas | 43 pagos (21 pendientes)", 64),
    ("Gustavo Primo",                "Bajaj OON 66H", "OON66H", datetime(2025, 4, 3), "moto 19 | 17 meses | 16320000 total | 68 cuotas | 68 pagos (0 pendientes)", 68),
    ("Erzik",                        "Bajaj OUN 39H", "OUN39H", datetime(2025, 8, 3), "moto 20 | 36 meses | 13968000 total | 36 cuotas | tarifa 384000/mes | ver fotos 37-38 para pagos adicionales", 36),
    ("Efrain",                       "Bajaj OTZ 01H", "OTZ01H", datetime(2025, 9, 2), "Amigo de Guillo | 21 meses | 20160000 total | 84 cuotas | subtotal cuaderno ~11.280.000 (≈47 cuotas) | algunos pagos son multiples por fila", 84),
    ("Gonzado Estherline",           "Bajaj OUB 81H", "OUB81H", datetime(2025, 9,23), "Cuñado de Katherine | del Choco | Tel: 3205570847 | 21 meses | 20160000 total | 84 cuotas | subtotal cuaderno 11.520.000 (48 cuotas) | 41 registradas fotos 38+39", 84),
    ("Ana Milena Juan Davit",        "Bajaj UPY 65H", "UPY65H", datetime(2026, 1,23), "Foto 40 | 25 pagos registrados | cuotas pendientes a confirmar", 64),
    ("Sr Luis",                      "Bajaj UYZ 32H", "UYZ32H", datetime(2026, 1, 5), "Foto 41 | 14 pagos registrados | nota: calibrar moto", 64),
    ("Manuel Alga",                  "Bajaj UYZ 39H", "UYZ39H", datetime(2026, 2,12), "Foto 42 | 18 meses | 72 cuotas | 18 pagos registrados", 72),
    ("Jorge Estrada",                 "Bajaj BJH 66I", "BJH66I", datetime(2026, 2,16), "Foto 43 | 64 cuotas | 18 meses | 16 pagos | GAP en mayo-2026 sin registro", 64),
    ("Donita Juan Camilo",           "Bajaj BJK 95I", "BJK95I", datetime(2026, 3, 5), "Foto 44 | 24 meses | 24 cuotas MENSUALES | tarifa 420.000/mes | total 10.080.000", 24),
    ("Alejandra",                    "Bajaj OKS 30I", "OKS30I", datetime(2026, 4,11), "Foto 45 | 21 meses | 84 cuotas | 13 pagos registrados", 84),
    ("Carlos",                       "Bajaj BLB 45I", "BLB45I", datetime(2026, 5,16), "Foto 46 | 18 meses | 72 cuotas | 4 pagos registrados | moto #29", 72),
    ("Francisco (Buñuelo)",          "Bajaj BLG 78I", "BLG78I", datetime(2026, 5, 3), "Foto 47 | 24 meses | 48 cuotas bisemanales | tarifa 190.000 | 4 pagos | moto #30", 48),
    ("Yesenia",                      "Bajaj HWO 62I", "HWO62I", datetime(2026, 7, 8), "Foto 48 | 18 meses | 72 cuotas | 3 pagos registrados | moto #31", 72),
    ("Alvania",                      "Bajaj HWD 78I", "HWD78I", datetime(2026, 7,10), "Foto 49 | 18 meses | 72 cuotas | 2 pagos registrados | moto #32", 72),
    ("Yurais Polo",                  "Bajaj HWM 11I", "HWM11I", datetime(2026, 8, 1), "Foto 50 | 18 meses | 72 cuotas | moto #33 | SIN PAGOS aun — cliente nuevo", 72),
]

# Pagos de Dorlys (57 pagos registrados en fotos 1 y 2)
PAGOS_DORLYS = [
    # Foto 1 (1–30)
    datetime(2024,12,26), datetime(2025, 1, 2), datetime(2025, 1, 9), datetime(2025, 1,16),
    datetime(2025, 1,23), datetime(2025, 2, 4), datetime(2025, 2,14), datetime(2025, 2,22),
    datetime(2025, 3, 1), datetime(2025, 3, 8), datetime(2025, 3,16), datetime(2025, 3,22),
    datetime(2025, 4, 3), datetime(2025, 4,12), datetime(2025, 4,26), datetime(2025, 5, 1),
    datetime(2025, 6, 5), datetime(2025, 6,18), datetime(2025, 6,24), datetime(2025, 7, 2),
    datetime(2025, 7,11), datetime(2025, 7,19), datetime(2025, 7,28), datetime(2025, 8, 6),
    datetime(2025, 8,14), datetime(2025, 8,22), datetime(2025, 8,31), datetime(2025, 9, 9),
    datetime(2025, 9,18), datetime(2025, 9,26),
    # Foto 2 (31–57)
    datetime(2025,10, 5), datetime(2025,10,13), datetime(2025,10,24), datetime(2025,11, 2),
    datetime(2025,11,12), datetime(2025,11,21), datetime(2025,11,29), datetime(2025,12, 8),
    datetime(2025,12,16), datetime(2025,12,25), datetime(2026, 1, 8), datetime(2026, 1,14),
    datetime(2026, 1,29), datetime(2026, 2, 3), datetime(2026, 2,10), datetime(2026, 2,18),
    datetime(2026, 2,28),
    # GAP: mar/2026 – abr/2026 sin registro en el cuaderno (verificar con tío)
    datetime(2026, 4,22), datetime(2026, 5, 2), datetime(2026, 5,10), datetime(2026, 5,19),
    datetime(2026, 5,27), datetime(2026, 6, 6), datetime(2026, 6,16), datetime(2026, 6,26),
    datetime(2026, 7, 7), datetime(2026, 7,16),
]

# Pagos de Juan Andres De Arco (63 pagos registrados en fotos 5 y 6 — falta 1 cuota)
PAGOS_JUAN_ANDRES = [
    # Foto 5 (1–30) — ene 2025 a ago 2025
    datetime(2025, 1, 5), datetime(2025, 1,13), datetime(2025, 1,21), datetime(2025, 1,29),
    datetime(2025, 2, 6), datetime(2025, 2,14), datetime(2025, 2,22), datetime(2025, 3, 2),
    datetime(2025, 3,10), datetime(2025, 3,18), datetime(2025, 3,26), datetime(2025, 4, 4),
    datetime(2025, 4,12), datetime(2025, 4,20), datetime(2025, 4,28), datetime(2025, 5, 6),
    datetime(2025, 5,14), datetime(2025, 5,22), datetime(2025, 5,30), datetime(2025, 6,10),
    datetime(2025, 6,16), datetime(2025, 6,25), datetime(2025, 7, 3), datetime(2025, 7,10),
    datetime(2025, 7,18), datetime(2025, 7,26), datetime(2025, 8, 3), datetime(2025, 8,11),
    datetime(2025, 8,20), datetime(2025, 8,28),
    # Foto 6 (31–63) — sep 2025 a jul 2026
    datetime(2025, 9, 8), datetime(2025, 9,13), datetime(2025, 9,21),
    datetime(2025,10, 6), datetime(2025,10,15), datetime(2025,10,23),
    datetime(2025,11, 5), datetime(2025,11,14), datetime(2025,11,24),
    datetime(2025,12, 2), datetime(2025,12,13), datetime(2025,12,19), datetime(2025,12,29),
    datetime(2026, 1,13), datetime(2026, 1,20),
    datetime(2026, 2, 2), datetime(2026, 2, 9), datetime(2026, 2,15), datetime(2026, 2,25),
    datetime(2026, 3, 3), datetime(2026, 3,10), datetime(2026, 3,17), datetime(2026, 3,29),
    datetime(2026, 4, 5), datetime(2026, 4,11), datetime(2026, 4,20), datetime(2026, 4,29),
    datetime(2026, 5,12),
    datetime(2026, 6, 2), datetime(2026, 6, 9),
    datetime(2026, 6,22), datetime(2026, 7, 8), datetime(2026, 7,22),
]

# Pagos de Jorge Luis Reyes (64 pagos registrados en fotos 7 y 8 — CONTRATO FINALIZADO)
PAGOS_JORGE_LUIS = [
    # Foto 7 (1–30) — ene 2025 a ago 2025
    datetime(2025, 1,10), datetime(2025, 1,18), datetime(2025, 1,26),
    datetime(2025, 2, 3), datetime(2025, 2,12), datetime(2025, 2,20), datetime(2025, 2,28),
    datetime(2025, 3, 8), datetime(2025, 3,16), datetime(2025, 3,27),
    datetime(2025, 4, 1), datetime(2025, 4, 9), datetime(2025, 4,17), datetime(2025, 4,25),
    datetime(2025, 5, 3), datetime(2025, 5,11), datetime(2025, 5,17), datetime(2025, 5,22),
    datetime(2025, 6, 4), datetime(2025, 6,12), datetime(2025, 6,20), datetime(2025, 6,28),
    datetime(2025, 7, 6), datetime(2025, 7,14), datetime(2025, 7,22), datetime(2025, 7,30),
    datetime(2025, 8, 7), datetime(2025, 8,15), datetime(2025, 8,23), datetime(2025, 8,31),
    # Foto 8 (31–64) — sep 2025 a jun 2026
    datetime(2025, 9, 8), datetime(2025, 9,16), datetime(2025, 9,24),
    datetime(2025,10, 1), datetime(2025,10, 8), datetime(2025,10,16), datetime(2025,10,24),
    datetime(2025,11, 1), datetime(2025,11, 9), datetime(2025,11,17), datetime(2025,11,25),
    datetime(2025,12, 3), datetime(2025,12,11), datetime(2025,12,18), datetime(2025,12,27),
    datetime(2026, 1, 9), datetime(2026, 1,17), datetime(2026, 1,25),
    datetime(2026, 2, 2), datetime(2026, 2,10), datetime(2026, 2,18), datetime(2026, 2,25),
    datetime(2026, 3, 5), datetime(2026, 3,13), datetime(2026, 3,21), datetime(2026, 3,29),
    datetime(2026, 4, 6), datetime(2026, 4, 8), datetime(2026, 4,18), datetime(2026, 4,24),
    datetime(2026, 5, 7), datetime(2026, 5,12), datetime(2026, 5,20), datetime(2026, 6, 3),
]

# Pagos de Duvon Enrique (63 pagos — fotos 10 y 11, 1 cuota pendiente)
PAGOS_DUVON = [
    # Foto 10 (1-30) — feb 2025 a oct 2025
    datetime(2025, 2,17), datetime(2025, 2,26),
    datetime(2025, 3,12), datetime(2025, 3,19), datetime(2025, 3,26),
    datetime(2025, 4, 8), datetime(2025, 4, 9),
    datetime(2025, 4,16), datetime(2025, 4,23),
    datetime(2025, 5, 1), datetime(2025, 5, 7), datetime(2025, 5,15),
    datetime(2025, 5,21), datetime(2025, 5,28),
    datetime(2025, 6, 4), datetime(2025, 6,10), datetime(2025, 6,19), datetime(2025, 6,25),
    datetime(2025, 7, 3), datetime(2025, 7,11), datetime(2025, 7,16),
    datetime(2025, 7,24), datetime(2025, 7,30),
    datetime(2025, 8, 7), datetime(2025, 8,14), datetime(2025, 8,21), datetime(2025, 8,28),
    datetime(2025, 9,25),
    datetime(2025,10, 1), datetime(2025,10, 8),
    # Foto 11 (31-63) — oct 2025 a jul 2026
    datetime(2025,10,15), datetime(2025,10,22), datetime(2025,10,29),
    datetime(2025,11, 5), datetime(2025,11,14), datetime(2025,11,19), datetime(2025,11,26),
    datetime(2025,12, 3), datetime(2025,12,10), datetime(2025,12,17),
    datetime(2025,12,24), datetime(2025,12,31),
    datetime(2026, 1, 9), datetime(2026, 1,21), datetime(2026, 1,29),
    datetime(2026, 2,19), datetime(2026, 2,27),
    datetime(2026, 3, 4), datetime(2026, 3,13), datetime(2026, 3,21), datetime(2026, 3,26),
    datetime(2026, 4, 2), datetime(2026, 4,16), datetime(2026, 4,20),
    datetime(2026, 5, 7), datetime(2026, 5,16), datetime(2026, 5,23), datetime(2026, 5,30),
    datetime(2026, 6, 7), datetime(2026, 6,20),
    datetime(2026, 7, 2), datetime(2026, 7,10), datetime(2026, 7,22),
]

# Pagos de Luis Gabriel Polo (24 pagos — moto repuesta dic 2025, columna 2 de foto 9)
# Moto anterior OYL39H fue robada; solo se registran pagos de la moto nueva KOM89H
PAGOS_LUIS_GABRIEL = [
    datetime(2025,12,23),
    datetime(2026, 1, 4), datetime(2026, 1, 8), datetime(2026, 1,15),
    datetime(2026, 1,22), datetime(2026, 1,29),
    datetime(2026, 2, 5), datetime(2026, 2,12), datetime(2026, 2,27),
    datetime(2026, 3, 5), datetime(2026, 3,12), datetime(2026, 3,19), datetime(2026, 3,26),
    datetime(2026, 4,10), datetime(2026, 4,15), datetime(2026, 4,18), datetime(2026, 4,25),
    datetime(2026, 5, 8), datetime(2026, 5,14), datetime(2026, 5,21), datetime(2026, 5,31),
    datetime(2026, 6, 6), datetime(2026, 6,16), datetime(2026, 6,23),
]

# Pagos de Osmed Brenda (54 pagos — fotos 16 y 17, OSZ46H, 64 cuotas)
# Último pago: 30/06/2026 — 21 días sin pagar al 28/07/2026 → atraso_alto
PAGOS_OSMED = [
    # Foto 16 (1-30) — may 2025 a dic 2025
    datetime(2025, 5,15), datetime(2025, 5,22), datetime(2025, 5,29),
    datetime(2025, 6, 5), datetime(2025, 6,12), datetime(2025, 6,20), datetime(2025, 6,27),
    datetime(2025, 7, 4), datetime(2025, 7,11), datetime(2025, 7,18), datetime(2025, 7,25),
    datetime(2025, 8, 1), datetime(2025, 8, 9), datetime(2025, 8,16),
    datetime(2025, 8,23), datetime(2025, 8,31),
    datetime(2025, 9, 7), datetime(2025, 9,14), datetime(2025, 9,21), datetime(2025, 9,27),
    datetime(2025,10, 4), datetime(2025,10,11), datetime(2025,10,18), datetime(2025,10,25),
    datetime(2025,11, 1), datetime(2025,11, 9), datetime(2025,11,20), datetime(2025,11,28),
    datetime(2025,12, 4), datetime(2025,12,11),
    # Foto 17 (31-54) — dic 2025 a jun 2026
    datetime(2025,12,18), datetime(2025,12,27),
    datetime(2026, 1,15), datetime(2026, 1,23), datetime(2026, 1,30),
    datetime(2026, 2, 8), datetime(2026, 2,18), datetime(2026, 2,25),
    datetime(2026, 3, 4), datetime(2026, 3,11), datetime(2026, 3,18), datetime(2026, 3,25),
    datetime(2026, 4, 1), datetime(2026, 4,11), datetime(2026, 4,18), datetime(2026, 4,26),
    datetime(2026, 5, 3), datetime(2026, 5,10), datetime(2026, 5,21), datetime(2026, 5,29),
    datetime(2026, 6, 5), datetime(2026, 6,13), datetime(2026, 6,22), datetime(2026, 6,30),
]

# Pagos de Jesus Morales Yuliana Olivares (52 pagos — fotos 14 y 15, KFK1AH, 72 cuotas)
PAGOS_JESUS = [
    # Foto 14 (1-30) — may 2025 a ene 2026
    datetime(2025, 5, 8), datetime(2025, 5,16), datetime(2025, 5,24),
    datetime(2025, 6, 1), datetime(2025, 6, 9), datetime(2025, 6,16), datetime(2025, 6,25),
    datetime(2025, 7, 3), datetime(2025, 7,11), datetime(2025, 7,19), datetime(2025, 7,27),
    datetime(2025, 8, 4), datetime(2025, 8,12), datetime(2025, 8,21), datetime(2025, 8,29),
    datetime(2025, 9, 6), datetime(2025, 9,15), datetime(2025, 9,23),
    datetime(2025,10, 1), datetime(2025,10,10), datetime(2025,10,18), datetime(2025,10,26),
    datetime(2025,11, 3), datetime(2025,11,11), datetime(2025,11,19), datetime(2025,11,27),
    datetime(2025,12, 5), datetime(2025,12,13), datetime(2025,12,22),
    datetime(2026, 1, 6),
    # Foto 15 (31-52) — ene 2026 a jul 2026
    datetime(2026, 1,14), datetime(2026, 1,23),
    datetime(2026, 2, 1), datetime(2026, 2,10), datetime(2026, 2,18), datetime(2026, 2,27),
    datetime(2026, 3, 8), datetime(2026, 3,18), datetime(2026, 3,26),
    datetime(2026, 4, 6), datetime(2026, 4,14), datetime(2026, 4,23),
    datetime(2026, 5, 2), datetime(2026, 5,10), datetime(2026, 5,18), datetime(2026, 5,26),
    datetime(2026, 6, 3), datetime(2026, 6,12), datetime(2026, 6,21), datetime(2026, 6,30),
    datetime(2026, 7, 8), datetime(2026, 7,14),
]

# Pagos de Wilmer Ivis (59 pagos — fotos 12 y 13, KEW23H, 64 cuotas)
PAGOS_WILMER = [
    # Foto 12 (1-30) — abr 2025 a nov 2025
    datetime(2025, 4, 4), datetime(2025, 4,12), datetime(2025, 4,20), datetime(2025, 4,28),
    datetime(2025, 5, 6), datetime(2025, 5,19), datetime(2025, 5,26), datetime(2025, 5,30),
    datetime(2025, 6, 7), datetime(2025, 6,16), datetime(2025, 6,23),
    datetime(2025, 7, 2), datetime(2025, 7,11), datetime(2025, 7,19), datetime(2025, 7,27),
    datetime(2025, 8, 4), datetime(2025, 8,12), datetime(2025, 8,20), datetime(2025, 8,28),
    datetime(2025, 9, 5), datetime(2025, 9,13), datetime(2025, 9,21), datetime(2025, 9,29),
    datetime(2025,10, 7), datetime(2025,10,13), datetime(2025,10,21), datetime(2025,10,30),
    datetime(2025,11, 8), datetime(2025,11,15), datetime(2025,11,23),
    # Foto 13 (31-59) — dic 2025 a jul 2026
    datetime(2025,12, 1), datetime(2025,12, 9), datetime(2025,12,17), datetime(2025,12,25),
    datetime(2026, 1, 7), datetime(2026, 1,15), datetime(2026, 1,23),
    datetime(2026, 2, 1), datetime(2026, 2, 9), datetime(2026, 2,17), datetime(2026, 2,25),
    datetime(2026, 3, 5), datetime(2026, 3,13), datetime(2026, 3,21), datetime(2026, 3,29),
    datetime(2026, 4, 7), datetime(2026, 4,15), datetime(2026, 4,23),
    datetime(2026, 5, 2), datetime(2026, 5,10), datetime(2026, 5,18), datetime(2026, 5,26),
    datetime(2026, 6, 3), datetime(2026, 6,11), datetime(2026, 6,19), datetime(2026, 6,28),
    datetime(2026, 7, 5), datetime(2026, 7,14), datetime(2026, 7,21),
]

# Pagos de Chacal Caterine (50 pagos — fotos 18-19, OSZ43H, 72 cuotas)
PAGOS_CHACAL = [
    # Foto 18 (1-30) — may 2025 a ene 2026
    datetime(2025, 5,13), datetime(2025, 5,25), datetime(2025, 6, 2), datetime(2025, 6,10),
    datetime(2025, 6,19), datetime(2025, 6,28), datetime(2025, 7, 6), datetime(2025, 7,15),
    datetime(2025, 7,24), datetime(2025, 8, 5), datetime(2025, 8,13), datetime(2025, 8,21),
    datetime(2025, 8,30), datetime(2025, 9, 7), datetime(2025, 9,15), datetime(2025, 9,22),
    datetime(2025, 9,28), datetime(2025,10, 6), datetime(2025,10,13), datetime(2025,10,21),
    datetime(2025,10,30), datetime(2025,11, 6), datetime(2025,11,18), datetime(2025,11,24),
    datetime(2025,12, 2), datetime(2025,12,14), datetime(2025,12,20),
    datetime(2026, 1, 6), datetime(2026, 1,15), datetime(2026, 1,24),
    # Foto 19 (31-50) — feb 2026 a jul 2026
    datetime(2026, 2, 2), datetime(2026, 2,10), datetime(2026, 2,18), datetime(2026, 2,26),
    datetime(2026, 3, 6), datetime(2026, 3,14), datetime(2026, 3,23), datetime(2026, 4, 1),
    datetime(2026, 4,10), datetime(2026, 4,19), datetime(2026, 4,29), datetime(2026, 5, 7),
    datetime(2026, 5,17), datetime(2026, 5,25), datetime(2026, 6, 3), datetime(2026, 6,12),
    datetime(2026, 6,22), datetime(2026, 6,30), datetime(2026, 7, 8), datetime(2026, 7,17),
]

# Pagos de Wilian Junior Darlis Esther (34 pagos — fotos 20-21, OTE02H, 60 cuotas)
PAGOS_WILIAN_JR = [
    # Foto 20 (1-30) — may 2025 a may 2026 (gap ago-oct 2025)
    datetime(2025, 5,22), datetime(2025, 5,31), datetime(2025, 6, 5),
    datetime(2025, 6,13), datetime(2025, 6,20), datetime(2025, 6,27), datetime(2025, 7, 4),
    datetime(2025, 7,30),
    # GAP agosto-octubre 2025
    datetime(2025,11,13), datetime(2025,11,22), datetime(2025,11,29),
    datetime(2025,12, 7), datetime(2025,12,14), datetime(2025,12,23),
    datetime(2026, 1, 6), datetime(2026, 1,13), datetime(2026, 1,21),
    datetime(2026, 2, 1), datetime(2026, 2, 9), datetime(2026, 2,18), datetime(2026, 2,26),
    datetime(2026, 3, 6), datetime(2026, 3,16), datetime(2026, 3,26),
    datetime(2026, 4, 3), datetime(2026, 4, 9), datetime(2026, 4,18), datetime(2026, 4,27),
    datetime(2026, 5, 6), datetime(2026, 5,16),
    # Foto 21 (31-34) — may-jun 2026
    datetime(2026, 5,22), datetime(2026, 5,29), datetime(2026, 6, 8), datetime(2026, 6,16),
]

# Pagos de Darwin Yaimis Ysabel (54 pagos — fotos 22-23, OTE80H, 64 cuotas)
PAGOS_DARWIN = [
    # Foto 22 (1-39) — may 2025 a mar 2026
    datetime(2025, 5,25), datetime(2025, 6, 3), datetime(2025, 6,10), datetime(2025, 6,16),
    datetime(2025, 6,22), datetime(2025, 6,28), datetime(2025, 7, 5), datetime(2025, 7,13),
    datetime(2025, 7,21), datetime(2025, 7,28), datetime(2025, 8, 4), datetime(2025, 8,18),
    datetime(2025, 8,25), datetime(2025, 9, 4), datetime(2025, 9,11), datetime(2025, 9,18),
    datetime(2025, 9,25), datetime(2025,10, 3), datetime(2025,10,11), datetime(2025,10,16),
    datetime(2025,10,23), datetime(2025,10,30), datetime(2025,11, 4), datetime(2025,11,11),
    datetime(2025,11,15), datetime(2025,11,25), datetime(2025,12, 1), datetime(2025,12, 8),
    datetime(2025,12,15), datetime(2025,12,26),
    datetime(2026, 1, 6), datetime(2026, 1,14), datetime(2026, 1,21), datetime(2026, 1,27),
    datetime(2026, 2, 4), datetime(2026, 2,12), datetime(2026, 2,20), datetime(2026, 2,27),
    datetime(2026, 3, 8),
    # Foto 23 (40-54) — mar-jul 2026
    datetime(2026, 3,17), datetime(2026, 3,23), datetime(2026, 3,29),
    datetime(2026, 4, 9), datetime(2026, 4,21), datetime(2026, 4,29),
    datetime(2026, 5, 9), datetime(2026, 5,19), datetime(2026, 5,29),
    datetime(2026, 6, 6), datetime(2026, 6,12), datetime(2026, 6,21),
    datetime(2026, 6,29), datetime(2026, 7, 9), datetime(2026, 7,17),
]

# Pagos de Kevin Rodriguez (29 pagos bisemanales — foto 24, OXO04H, 84 cuotas)
PAGOS_KEVIN = [
    # Foto 24 (1-29) — may 2025 a jun 2026
    datetime(2025, 5,27), datetime(2025, 6,10), datetime(2025, 6,24),
    datetime(2025, 7, 8), datetime(2025, 7,22), datetime(2025, 8, 5),
    datetime(2025, 8,19), datetime(2025, 9, 2), datetime(2025, 9,16),
    datetime(2025, 9,30), datetime(2025,10,14), datetime(2025,10,28),
    datetime(2025,11,11), datetime(2025,11,25), datetime(2025,12, 9),
    datetime(2025,12,23), datetime(2026, 1, 6), datetime(2026, 1,20),
    datetime(2026, 2, 3), datetime(2026, 2,17), datetime(2026, 3, 3),
    datetime(2026, 3,17), datetime(2026, 3,31), datetime(2026, 4,14),
    datetime(2026, 4,28), datetime(2026, 5,12), datetime(2026, 5,26),
    datetime(2026, 6, 9), datetime(2026, 6,28),
]

# Pagos de Nabro Maria Jesse (51 pagos — fotos 25-26, OTE68H, 64 cuotas)
PAGOS_NABRO = [
    # Foto 25 (1-44) — abr 2025 a mar 2026
    datetime(2025, 4,22), datetime(2025, 4,30), datetime(2025, 5, 8), datetime(2025, 5,16),
    datetime(2025, 5,25), datetime(2025, 6, 2), datetime(2025, 6,11), datetime(2025, 6,18),
    datetime(2025, 6,25), datetime(2025, 7, 2), datetime(2025, 7,11), datetime(2025, 7,19),
    datetime(2025, 7,26), datetime(2025, 8, 2), datetime(2025, 8,10), datetime(2025, 8,18),
    datetime(2025, 8,26), datetime(2025, 9, 3), datetime(2025, 9,11), datetime(2025, 9,18),
    datetime(2025, 9,25), datetime(2025,10, 2), datetime(2025,10,10), datetime(2025,10,18),
    datetime(2025,10,25), datetime(2025,11, 1), datetime(2025,11, 9), datetime(2025,11,17),
    datetime(2025,11,25), datetime(2025,12, 3), datetime(2025,12,11), datetime(2025,12,19),
    datetime(2025,12,27), datetime(2026, 1, 4), datetime(2026, 1,12), datetime(2026, 1,20),
    datetime(2026, 1,28), datetime(2026, 2, 5), datetime(2026, 2,13), datetime(2026, 2,21),
    datetime(2026, 3, 1), datetime(2026, 3, 9), datetime(2026, 3,17), datetime(2026, 3,25),
    # Foto 26 (45-51) — abr-jul 2026
    datetime(2026, 4,21), datetime(2026, 5, 1), datetime(2026, 5,22),
    datetime(2026, 6, 8), datetime(2026, 6,22), datetime(2026, 7, 2), datetime(2026, 7,20),
]

# Pagos de Luis Katherin Estor (54 pagos — fotos 27-28, OTG 82H, 69 cuotas)
# Foto 28 = pagos 1-30 (jun 2025 – ene 2026); foto 27 = pagos 31-54 (ene-jul 2026)
PAGOS_LUIS_KATHERIN = [
    # Foto 28 (1-30) — jun 2025 a ene 2026
    datetime(2025, 6, 4), datetime(2025, 6,11), datetime(2025, 6,18), datetime(2025, 6,25),
    datetime(2025, 7, 2), datetime(2025, 7, 9), datetime(2025, 7,16), datetime(2025, 7,23),
    datetime(2025, 7,30), datetime(2025, 8, 6), datetime(2025, 8,13), datetime(2025, 8,20),
    datetime(2025, 8,27), datetime(2025, 9, 3), datetime(2025, 9,10),
    datetime(2025,10, 1), datetime(2025,10,17), datetime(2025,10,20), datetime(2025,10,24),
    datetime(2025,10,25), datetime(2025,10,29),
    datetime(2025,11, 7), datetime(2025,11,17), datetime(2025,11,17),
    datetime(2025,11,26), datetime(2025,11,26), datetime(2025,11,26), datetime(2025,11,26),
    datetime(2025,12,24), datetime(2026, 1,14),
    # Foto 27 (31-54) — ene 2026 a jul 2026
    datetime(2026, 1,29), datetime(2026, 2, 8), datetime(2026, 2,16),
    datetime(2026, 3,15), datetime(2026, 3,15), datetime(2026, 3,19), datetime(2026, 3,23),
    datetime(2026, 4, 1), datetime(2026, 4, 1), datetime(2026, 4, 9), datetime(2026, 4,16),
    datetime(2026, 4,22), datetime(2026, 4,30),
    datetime(2026, 5,13), datetime(2026, 5,21), datetime(2026, 5,28),
    datetime(2026, 6, 9), datetime(2026, 6,11), datetime(2026, 6,18), datetime(2026, 6,25),
    datetime(2026, 7, 3), datetime(2026, 7,10), datetime(2026, 7,16), datetime(2026, 7,23),
]

# Pagos de Elkin (58 pagos registrados en fotos 3 y 4)
PAGOS_ELKIN = [
    # Foto 3 (1–30)
    datetime(2024,12,26), datetime(2025, 1, 2), datetime(2025, 1, 9), datetime(2025, 1,16),
    datetime(2025, 1,27), datetime(2025, 2, 4), datetime(2025, 2,22), datetime(2025, 3, 2),
    datetime(2025, 3,11), datetime(2025, 3,14), datetime(2025, 3,27), datetime(2025, 4, 4),
    datetime(2025, 4,12), datetime(2025, 4,20), datetime(2025, 4,28), datetime(2025, 5, 6),
    datetime(2025, 5,16), datetime(2025, 5,24), datetime(2025, 6, 3), datetime(2025, 6, 8),
    datetime(2025, 6,14), datetime(2025, 6,30), datetime(2025, 7, 9), datetime(2025, 7,18),
    datetime(2025, 7,29), datetime(2025, 8, 7), datetime(2025, 8,16), datetime(2025, 8,24),
    datetime(2025, 9, 2), datetime(2025, 9,10),
    # Foto 4 (31–58)
    datetime(2025, 9,18), datetime(2025, 9,26), datetime(2025,10, 4), datetime(2025,10,14),
    datetime(2025,10,21), datetime(2025,10,31), datetime(2025,11, 9), datetime(2025,11,19),
    datetime(2025,11,26), datetime(2025,12, 6), datetime(2025,12,14), datetime(2025,12,23),
    datetime(2026, 1, 7), datetime(2026, 1,16), datetime(2026, 1,24), datetime(2026, 2, 2),
    datetime(2026, 2,13), datetime(2026, 2,24), datetime(2026, 3, 2), datetime(2026, 3,15),
    datetime(2026, 3,30), datetime(2026, 4,18), datetime(2026, 4,28), datetime(2026, 5,16),
    datetime(2026, 5,20), datetime(2026, 6,15), datetime(2026, 7, 1), datetime(2026, 7,21),
]

# Pagos de Lucho Laura Nosu (51 pagos — fotos 29-30, OTF59H, 64 cuotas)
PAGOS_LUCHO = [
    # Foto 29 (1-30) — may 2025 a ene 2026
    datetime(2025, 5,31), datetime(2025, 6, 8), datetime(2025, 6,16), datetime(2025, 6,24),
    datetime(2025, 7, 2), datetime(2025, 7,10), datetime(2025, 7,12), datetime(2025, 7,26),
    datetime(2025, 8, 2), datetime(2025, 8,10), datetime(2025, 8,18), datetime(2025, 8,26),
    datetime(2025, 9, 3), datetime(2025, 9,11), datetime(2025, 9,19), datetime(2025, 9,27),
    datetime(2025,10, 5), datetime(2025,10,14), datetime(2025,10,22), datetime(2025,10,30),
    datetime(2025,11, 7), datetime(2025,11, 8), datetime(2025,11,23), datetime(2025,11,30),
    datetime(2025,12,13), datetime(2025,12,18), datetime(2025,12,25),
    datetime(2026, 1, 8), datetime(2026, 1,17), datetime(2026, 1,25),
    # Foto 30 (31-51) — feb 2026 a jul 2026
    datetime(2026, 2, 4), datetime(2026, 2,12), datetime(2026, 2,20), datetime(2026, 2,28),
    datetime(2026, 3, 8), datetime(2026, 3,17), datetime(2026, 3,26),
    datetime(2026, 4, 4), datetime(2026, 4,12), datetime(2026, 4,20), datetime(2026, 4,28),
    datetime(2026, 5, 6), datetime(2026, 5,15), datetime(2026, 5,23), datetime(2026, 5,31),
    datetime(2026, 6, 9), datetime(2026, 6,18), datetime(2026, 6,27),
    datetime(2026, 7, 5), datetime(2026, 7,13), datetime(2026, 7,21),
]

# Pagos de Roberto (43 pagos — foto 31, OTR 29H, 64 cuotas)
# Formato de dos columnas en la misma página:
#   izquierda = pagos jun 2025 – ene 2026
#   derecha   = pagos feb 2026 – jul 2026
# NOTA: fechas de las filas 4-6 (ago-sep 2025) aproximadas — verificar con el tío
PAGOS_ROBERTO = [
    # Foto 31 columna izquierda (1-23) — jun 2025 a ene 2026
    datetime(2025, 6,24), datetime(2025, 7, 1), datetime(2025, 7,28),
    datetime(2025, 8, 8),   # aprox
    datetime(2025, 8,24),   # aprox
    datetime(2025, 9, 5),
    datetime(2025, 9,12), datetime(2025, 9,18), datetime(2025, 9,26),
    datetime(2025,10, 6), datetime(2025,10,14), datetime(2025,10,18), datetime(2025,10,27),
    datetime(2025,11, 2), datetime(2025,11,13), datetime(2025,11,24), datetime(2025,11,29),
    datetime(2025,12,10), datetime(2025,12,16), datetime(2025,12,23), datetime(2025,12,29),
    datetime(2026, 1,10), datetime(2026, 1,23),
    # Foto 31 columna derecha (1-20) — feb 2026 a jul 2026
    datetime(2026, 2,17), datetime(2026, 2,24),
    datetime(2026, 3, 3), datetime(2026, 3,10), datetime(2026, 3,17), datetime(2026, 3,24),
    datetime(2026, 4, 1), datetime(2026, 4, 8), datetime(2026, 4,16), datetime(2026, 4,25),
    datetime(2026, 5, 3), datetime(2026, 5,12), datetime(2026, 5,18), datetime(2026, 5,31),
    datetime(2026, 6,10), datetime(2026, 6,15), datetime(2026, 6,23), datetime(2026, 6,28),
    datetime(2026, 7, 3), datetime(2026, 7,16),
]

# Pagos de Braillon (43 pagos — fotos 32 y 33, OUA 19H, 64 cuotas)
# Foto 33 está girada; fechas abr-jul 2026 son aproximadas — verificar con el tío
PAGOS_BRAILLON = [
    # Foto 32 (1-30) — jul 2025 a abr 2026
    datetime(2025, 7,17), datetime(2025, 7,24), datetime(2025, 7,31),
    datetime(2025, 8,16),
    datetime(2025, 9, 4), datetime(2025, 9,12), datetime(2025, 9,18), datetime(2025, 9,26),
    datetime(2025,10, 8), datetime(2025,10,12), datetime(2025,10,19), datetime(2025,10,27),
    datetime(2025,11, 4), datetime(2025,11,12), datetime(2025,11,21), datetime(2025,11,29),
    datetime(2025,12, 7), datetime(2025,12,16), datetime(2025,12,24),
    datetime(2026, 1, 8), datetime(2026, 1,16), datetime(2026, 1,24),
    datetime(2026, 2, 2), datetime(2026, 2,10), datetime(2026, 2,18), datetime(2026, 2,27),
    datetime(2026, 3, 8), datetime(2026, 3,16), datetime(2026, 3,29),
    datetime(2026, 4, 2),
    # Foto 33 (31-43) — abr 2026 a jul 2026 (foto girada, fechas aprox)
    datetime(2026, 4,14), datetime(2026, 4,19),
    datetime(2026, 4,29),   # aprox
    datetime(2026, 5, 6), datetime(2026, 5,12), datetime(2026, 5,22),
    datetime(2026, 5,29),   # aprox
    datetime(2026, 6, 8), datetime(2026, 6,15), datetime(2026, 6,22),
    datetime(2026, 6,29),   # aprox
    datetime(2026, 7, 9), datetime(2026, 7,16),
]

# Pagos de Gustavo Primo (68 pagos — fotos 34 y 35, moto 19)
PAGOS_GUSTAVO = [
    # Foto 34 (1-52) — abr 2025 a mar 2026 (foto girada, fechas aprox semanales)
    datetime(2025, 4, 3), datetime(2025, 4,10), datetime(2025, 4,17), datetime(2025, 4,24),
    datetime(2025, 5, 1), datetime(2025, 5, 8), datetime(2025, 5,15), datetime(2025, 5,22),
    datetime(2025, 5,29),
    datetime(2025, 6, 5), datetime(2025, 6,12), datetime(2025, 6,19), datetime(2025, 6,26),
    datetime(2025, 7, 3), datetime(2025, 7,10), datetime(2025, 7,17), datetime(2025, 7,24),
    datetime(2025, 7,31),
    datetime(2025, 8, 7), datetime(2025, 8,14), datetime(2025, 8,21), datetime(2025, 8,28),
    datetime(2025, 9, 4), datetime(2025, 9,11), datetime(2025, 9,18), datetime(2025, 9,25),
    datetime(2025,10, 2), datetime(2025,10, 9), datetime(2025,10,16), datetime(2025,10,23),
    datetime(2025,10,30),
    datetime(2025,11, 6), datetime(2025,11,13), datetime(2025,11,20), datetime(2025,11,27),
    datetime(2025,12, 4), datetime(2025,12,11), datetime(2025,12,18), datetime(2025,12,25),
    datetime(2026, 1, 1), datetime(2026, 1, 8), datetime(2026, 1,15), datetime(2026, 1,22),
    datetime(2026, 1,29),
    datetime(2026, 2, 5), datetime(2026, 2,12), datetime(2026, 2,19), datetime(2026, 2,26),
    datetime(2026, 3, 5), datetime(2026, 3,12), datetime(2026, 3,19), datetime(2026, 3,26),
    # Foto 35 (53-68) — abr 2026 a jul 2026
    datetime(2026, 4, 1), datetime(2026, 4, 8), datetime(2026, 4,15), datetime(2026, 4,22),
    datetime(2026, 4,29),
    datetime(2026, 5, 6), datetime(2026, 5,14), datetime(2026, 5,21), datetime(2026, 5,28),
    datetime(2026, 6, 4), datetime(2026, 6,11), datetime(2026, 6,18), datetime(2026, 6,25),
    datetime(2026, 7, 2), datetime(2026, 7, 9), datetime(2026, 7,16),
]

# Pagos de Efrain (foto 37 — 22 fechas; subtotal cuaderno ≈11.280.000 = 47 cuotas reales)
# Varios renglones tienen pagos múltiples (480k/600k = 2-3 cuotas). Fecha registrada = 1 entrada.
PAGOS_EFRAIN = [
    # Foto 37 — sep 2025 a jul 2026
    datetime(2025, 9, 2), datetime(2025, 9,19), datetime(2025, 9,26),
    datetime(2025,10, 3), datetime(2025,10,10), datetime(2025,10,24),
    datetime(2025,11, 4), datetime(2025,11,11), datetime(2025,11,23),
    datetime(2025,12, 5), datetime(2025,12,12), datetime(2025,12,23),
    datetime(2026, 1,12), datetime(2026, 1,31),
    datetime(2026, 2,12),
    datetime(2026, 3, 7), datetime(2026, 3,20),
    datetime(2026, 4,10),
    datetime(2026, 5, 1), datetime(2026, 5,20),
    datetime(2026, 6,25),
    datetime(2026, 7,22),
]

# Pagos de Gonzado Estherline (fotos 38 + 39 — 41 fechas; subtotal cuaderno 11.520.000 = 48 cuotas)
PAGOS_GONZADO = [
    # Foto 38 (1-30) — sep 2025 a abr 2026, pagos semanales aprox
    datetime(2025, 9,23), datetime(2025, 9,30),
    datetime(2025,10, 7), datetime(2025,10,14), datetime(2025,10,21), datetime(2025,10,28),
    datetime(2025,11, 4), datetime(2025,11,11), datetime(2025,11,18), datetime(2025,11,25),
    datetime(2025,12, 2), datetime(2025,12, 9), datetime(2025,12,16), datetime(2025,12,23), datetime(2025,12,30),
    datetime(2026, 1, 6), datetime(2026, 1,13), datetime(2026, 1,20), datetime(2026, 1,27),
    datetime(2026, 2, 3), datetime(2026, 2,10), datetime(2026, 2,17), datetime(2026, 2,24),
    datetime(2026, 3, 3), datetime(2026, 3,10), datetime(2026, 3,17), datetime(2026, 3,24), datetime(2026, 3,31),
    datetime(2026, 4, 7), datetime(2026, 4,14),
    # Foto 39 (31-41) — may 2026 a ago 2026
    datetime(2026, 5, 1), datetime(2026, 5,23), datetime(2026, 5,30),
    datetime(2026, 6, 9), datetime(2026, 6,15), datetime(2026, 6,26),
    datetime(2026, 7, 7), datetime(2026, 7,12), datetime(2026, 7,19), datetime(2026, 7,26),
    datetime(2026, 8, 1),
]

# Pagos de Manuel Alga (18 pagos — foto 42)
PAGOS_MANUEL_ALGA = [
    # Foto 42 — feb 2026 a jul 2026
    datetime(2026, 2,12), datetime(2026, 2,24),
    datetime(2026, 3, 5), datetime(2026, 3, 9), datetime(2026, 3,16), datetime(2026, 3,25),
    datetime(2026, 4, 6), datetime(2026, 4,14), datetime(2026, 4,25),
    datetime(2026, 5, 4), datetime(2026, 5,19), datetime(2026, 5,29),
    datetime(2026, 6, 8), datetime(2026, 6,17), datetime(2026, 6,25),
    datetime(2026, 7, 5), datetime(2026, 7, 8), datetime(2026, 7,21),
]

# Pagos de Carlos Habite (16 pagos — foto 43, dos columnas de fechas)
PAGOS_CARLOS_HABITE = [
    # Columna izquierda — feb a abr 2026
    datetime(2026, 2,16), datetime(2026, 2,24),
    datetime(2026, 3, 1), datetime(2026, 3, 9), datetime(2026, 3,17), datetime(2026, 3,24),
    datetime(2026, 4, 3), datetime(2026, 4,14), datetime(2026, 4,21),
    # Columna derecha — jun a jul 2026 (sin registros en mayo)
    datetime(2026, 6, 1), datetime(2026, 6, 9), datetime(2026, 6,16), datetime(2026, 6,24),
    datetime(2026, 7, 1), datetime(2026, 7,10), datetime(2026, 7,21),
]

# Pagos de Ana Milena / Juan Davit (25 pagos — foto 40)
PAGOS_ANA_MILENA = [
    # Foto 40 — ene 2026 a jul 2026, pagos semanales aprox
    datetime(2026, 1,23), datetime(2026, 2, 1),
    datetime(2026, 2,11), datetime(2026, 2,18), datetime(2026, 2,26),
    datetime(2026, 3, 5), datetime(2026, 3,12), datetime(2026, 3,19), datetime(2026, 3,26),
    datetime(2026, 4, 2), datetime(2026, 4, 9), datetime(2026, 4,16), datetime(2026, 4,23), datetime(2026, 4,30),
    datetime(2026, 5, 7), datetime(2026, 5,14), datetime(2026, 5,21), datetime(2026, 5,28),
    datetime(2026, 6, 4), datetime(2026, 6,11), datetime(2026, 6,18), datetime(2026, 6,25),
    datetime(2026, 7, 2), datetime(2026, 7, 9), datetime(2026, 7,16),
]

# Pagos de Sr Luis (14 pagos — foto 41)
PAGOS_SR_LUIS = [
    # Foto 41 — ene 2026 a abr 2026, pagos semanales aprox
    datetime(2026, 1, 5), datetime(2026, 1,13), datetime(2026, 1,21), datetime(2026, 1,29),
    datetime(2026, 2, 6), datetime(2026, 2,14), datetime(2026, 2,22),
    datetime(2026, 3, 2), datetime(2026, 3,10), datetime(2026, 3,18), datetime(2026, 3,26),
    datetime(2026, 4, 3), datetime(2026, 4,11), datetime(2026, 4,19),
]

TARIFA_ERZIK       = 384_000  # mensual (≠ 240.000 semanal de otros clientes)
TARIFA_DONITA      = 420_000  # mensual (foto 44 — pago único mensual, 24 cuotas)
TARIFA_FRANCISCO   = 190_000  # bisemanal (foto 47 — 48 cuotas en 24 meses)

# Pagos de Carlos (4 pagos — foto 46, BLB 45I, 18 meses / 72 cuotas, moto #29)
PAGOS_CARLOS = [
    datetime(2026, 5,16), datetime(2026, 5,24), datetime(2026, 6, 4), datetime(2026, 6,18),
]

# Pagos de Francisco Buñuelo (4 pagos — foto 47, BLG 78I, 24 meses / 48 cuotas bisemanales)
PAGOS_FRANCISCO = [
    datetime(2026, 5, 3), datetime(2026, 5,15), datetime(2026, 5,30), datetime(2026, 7,15),
]

# Pagos de Yesenia (3 pagos — foto 48, HWO 62I, 18 meses / 72 cuotas, moto #31)
PAGOS_YESENIA = [
    datetime(2026, 7, 8), datetime(2026, 7,15), datetime(2026, 7,22),
]

# Pagos de Alvania (2 pagos — foto 49, HWD 78I, 18 meses / 72 cuotas, moto #32)
PAGOS_ALVANIA = [
    datetime(2026, 7,10), datetime(2026, 7,17),
]

# Pagos de Donita Juan Camilo (4 pagos — foto 44, 24 cuotas MENSUALES de 420.000)
PAGOS_DONITA_JUAN_CAMILO = [
    datetime(2026, 3, 5), datetime(2026, 4, 5), datetime(2026, 5, 5), datetime(2026, 6, 2),
]

# Pagos de Alejandra (13 pagos — foto 45, OKS 30I, 21 meses / 84 cuotas)
PAGOS_ALEJANDRA = [
    # Foto 45 — abr 2026 a jul 2026, pagos semanales aprox
    datetime(2026, 4,11), datetime(2026, 4,23),
    datetime(2026, 5, 2), datetime(2026, 5, 9), datetime(2026, 5,16), datetime(2026, 5,24),
    datetime(2026, 6, 1), datetime(2026, 6, 9), datetime(2026, 6,16), datetime(2026, 6,24),
    datetime(2026, 7, 1), datetime(2026, 7, 7), datetime(2026, 7,13),
]

# Pagos de Erzik (foto 36 — ver fotos 37-38 para pagos restantes hasta completar 36)
PAGOS_ERZIK = [
    # Foto 36 — ago 2025 a jul 2026 (pagos aprox mensuales, algunos meses doble pago)
    datetime(2025, 8, 3), datetime(2025, 9, 7), datetime(2025,10, 5),
    datetime(2025,11, 8), datetime(2025,12, 4),
    datetime(2026, 1, 9), datetime(2026, 1,31),  # dos pagos en enero
    datetime(2026, 2,10),
    datetime(2026, 3, 6), datetime(2026, 3,14),  # dos pagos en marzo
    datetime(2026, 4, 3),
    datetime(2026, 5, 3), datetime(2026, 5,19),  # dos pagos en mayo
    datetime(2026, 6, 5), datetime(2026, 6,19),  # dos pagos en junio
    datetime(2026, 7, 3), datetime(2026, 7,16),
]

# ---------------------------------------------------------------------------
# CREAR WORKBOOK
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()

# ---- Hoja Clientes ----
ws_cli = wb.active
ws_cli.title = "Clientes"

hdr_fill = PatternFill("solid", fgColor="211E1A")
hdr_font = Font(bold=True, color="F4EFE2")

cabeceras_cli = ["Nombre", "Moto", "Placa", "Inicio", "Observaciones", "Total Cuotas"]
ws_cli.append(cabeceras_cli)
for c, _ in enumerate(cabeceras_cli, 1):
    cell = ws_cli.cell(1, c)
    cell.fill = hdr_fill
    cell.font = hdr_font

for nombre, moto, placa, inicio, obs, total_cuotas in CLIENTES:
    ws_cli.append([nombre, moto, placa, inicio, obs, total_cuotas])

ws_cli.column_dimensions["A"].width = 24
ws_cli.column_dimensions["B"].width = 18
ws_cli.column_dimensions["C"].width = 10
ws_cli.column_dimensions["D"].width = 14
ws_cli.column_dimensions["E"].width = 44
ws_cli.column_dimensions["F"].width = 14

# ---- Hoja Registro Diario ----
ws_reg = wb.create_sheet("Registro Diario")

# 16 columnas para coincidir con los índices que usa el script Python:
# A=fecha B=cliente C=moto D=placa E-G=libre H=pago_diario
# I=pago_recibido J=medio_pago K=saldo L-O=libre P=observaciones
cabeceras_reg = [
    "Fecha", "Cliente", "Moto", "Placa", "", "", "",
    "Pago Diario", "Pago Recibido", "Medio Pago", "Saldo",
    "", "", "", "", "Observaciones"
]
ws_reg.append(cabeceras_reg)
for c, h in enumerate(cabeceras_reg, 1):
    cell = ws_reg.cell(1, c)
    if h:
        cell.fill = hdr_fill
        cell.font = hdr_font

# Función auxiliar para agregar filas
def agregar_pagos(nombre, moto, placa, pagos, tarifa=TARIFA):
    for fecha in pagos:
        row = [None] * 16
        row[0]  = fecha          # A: Fecha
        row[1]  = nombre         # B: Cliente
        row[2]  = moto           # C: Moto
        row[3]  = placa          # D: Placa
        row[7]  = tarifa         # H: Pago Diario
        row[8]  = tarifa         # I: Pago Recibido (se asume pagado según cuaderno)
        row[9]  = "efectivo"     # J: Medio Pago
        row[10] = None           # K: Saldo (el script lo calcula)
        ws_reg.append(row)

agregar_pagos("Dorlys Jose Julio",     "Bajaj EYO 114", "EYO114", PAGOS_DORLYS)
agregar_pagos("Elkin Espitia Payares", "Bajaj EYO 06H", "EYO06H", PAGOS_ELKIN)
agregar_pagos("Juan Andres De Arco",   "Bajaj KOO 16H", "KOO16H", PAGOS_JUAN_ANDRES)
agregar_pagos("Jorge Luis Reyes",      "Bajaj KOO B1H", "KOOB1H", PAGOS_JORGE_LUIS)
agregar_pagos("Luis Gabriel Polo",     "Bajaj KOM 89H", "KOM89H", PAGOS_LUIS_GABRIEL)
agregar_pagos("Duvon Enrique",         "Bajaj KEB 58H", "KEB58H", PAGOS_DUVON)
agregar_pagos("Wilmer Ivis",           "Bajaj KEW 23H", "KEW23H", PAGOS_WILMER)
agregar_pagos("Jesus Morales Yuliana Olivares", "Bajaj KFK 1AH", "KFK1AH", PAGOS_JESUS)
agregar_pagos("Osmed Brenda",                  "Bajaj OSZ 46H", "OSZ46H", PAGOS_OSMED)
agregar_pagos("Chacal Caterine",               "Bajaj OSZ 43H", "OSZ43H", PAGOS_CHACAL)
agregar_pagos("Wilian Junior Darlis Esther",   "Bajaj OTE 02H", "OTE02H", PAGOS_WILIAN_JR)
agregar_pagos("Darwin Yaimis Ysabel",          "Bajaj OTE 80H", "OTE80H", PAGOS_DARWIN)
agregar_pagos("Kevin Rodriguez",               "Bajaj OXO 04H", "OXO04H", PAGOS_KEVIN)
agregar_pagos("Nabro Maria Jesse",             "Bajaj OTE 68H", "OTE68H", PAGOS_NABRO)
agregar_pagos("Luis Katherin Estor",           "Bajaj OTG 82H", "OTG82H", PAGOS_LUIS_KATHERIN)
agregar_pagos("Lucho Laura Nosu",             "Bajaj OTF 59H", "OTF59H", PAGOS_LUCHO)
agregar_pagos("Roberto",                      "Bajaj OTR 29H", "OTR29H", PAGOS_ROBERTO)
agregar_pagos("Braillon",                     "Bajaj OUA 19H", "OUA19H", PAGOS_BRAILLON)
agregar_pagos("Gustavo Primo",                "Bajaj OON 66H", "OON66H", PAGOS_GUSTAVO)
agregar_pagos("Erzik",                        "Bajaj OUN 39H", "OUN39H", PAGOS_ERZIK, tarifa=TARIFA_ERZIK)
agregar_pagos("Efrain",                       "Bajaj OTZ 01H", "OTZ01H", PAGOS_EFRAIN)
agregar_pagos("Gonzado Estherline",           "Bajaj OUB 81H", "OUB81H", PAGOS_GONZADO)
agregar_pagos("Ana Milena Juan Davit",        "Bajaj UPY 65H", "UPY65H", PAGOS_ANA_MILENA)
agregar_pagos("Sr Luis",                      "Bajaj UYZ 32H", "UYZ32H", PAGOS_SR_LUIS)
agregar_pagos("Manuel Alga",                  "Bajaj UYZ 39H", "UYZ39H", PAGOS_MANUEL_ALGA)
agregar_pagos("Jorge Estrada",                "Bajaj BJH 66I", "BJH66I", PAGOS_CARLOS_HABITE)
agregar_pagos("Donita Juan Camilo",          "Bajaj BJK 95I", "BJK95I", PAGOS_DONITA_JUAN_CAMILO, tarifa=TARIFA_DONITA)
agregar_pagos("Alejandra",                   "Bajaj OKS 30I", "OKS30I", PAGOS_ALEJANDRA)
agregar_pagos("Carlos",                      "Bajaj BLB 45I", "BLB45I", PAGOS_CARLOS)
agregar_pagos("Francisco (Buñuelo)",         "Bajaj BLG 78I", "BLG78I", PAGOS_FRANCISCO, tarifa=TARIFA_FRANCISCO)
agregar_pagos("Yesenia",                     "Bajaj HWO 62I", "HWO62I", PAGOS_YESENIA)
agregar_pagos("Alvania",                     "Bajaj HWD 78I", "HWD78I", PAGOS_ALVANIA)

# Formato de fecha en columna A
for row in ws_reg.iter_rows(min_row=2):
    if row[0].value:
        row[0].number_format = "DD/MM/YYYY"

ws_reg.column_dimensions["A"].width = 13
ws_reg.column_dimensions["B"].width = 24
ws_reg.column_dimensions["C"].width = 16
ws_reg.column_dimensions["D"].width = 10
ws_reg.column_dimensions["H"].width = 14
ws_reg.column_dimensions["I"].width = 14
ws_reg.column_dimensions["J"].width = 12
ws_reg.column_dimensions["K"].width = 14
ws_reg.column_dimensions["P"].width = 36

# ─────────────────────────────────────────────────────────────────────────────
# HOJA: Simulador de Crédito
# ─────────────────────────────────────────────────────────────────────────────
ws_sim = wb.create_sheet("Simulador")

_thin = Side(style="thin")
_brd  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ── Título ──────────────────────────────────────────────────────────────────
ws_sim.merge_cells("A1:F1")
ws_sim.row_dimensions[1].height = 38
c = ws_sim["A1"]
c.value = "SIMULADOR DE CRÉDITO"
c.font  = Font(bold=True, size=16, color="F4EFE2")
c.fill  = PatternFill("solid", fgColor="211E1A")
c.alignment = Alignment(horizontal="center", vertical="center")

ws_sim.row_dimensions[2].height = 8  # spacer

# ── Parámetros de entrada ────────────────────────────────────────────────────
ws_sim.merge_cells("A3:F3")
ws_sim.row_dimensions[3].height = 24
c = ws_sim["A3"]
c.value = "  PARÁMETROS DE ENTRADA"
c.font  = Font(bold=True, size=11, color="F4EFE2")
c.fill  = PatternFill("solid", fgColor="3B3028")
c.alignment = Alignment(horizontal="left", vertical="center")

_param_rows = [
    (4, "Valor a Financiar  ($)",             None,      "← escribe el monto que vas a prestar"),
    (5, "Interés Mensual  (%)",               None,      "← ej: 3.5  significa  3.5 % mensual"),
    (6, "Interés Anual  (%)  —  automático",  "=B5*12",  "← se calcula solo"),
    (7, "Número de Cuotas  (meses)",          None,      "← ej: 12, 24, 36, 64…"),
]
for _row, _label, _formula, _hint in _param_rows:
    ws_sim.row_dimensions[_row].height = 24
    c = ws_sim[f"A{_row}"]
    c.value = _label
    c.font  = Font(bold=True, size=10)
    c.fill  = PatternFill("solid", fgColor="EAE7E4")
    c.border = _brd
    c.alignment = Alignment(horizontal="left", vertical="center")

    b = ws_sim[f"B{_row}"]
    if _formula:
        b.value = _formula
        b.fill  = PatternFill("solid", fgColor="D4EDDA")
    else:
        b.value = 0
        b.fill  = PatternFill("solid", fgColor="FFF3CD")  # amarillo = editable
    b.font   = Font(bold=True, size=11)
    b.border = _brd
    b.alignment = Alignment(horizontal="center", vertical="center")
    b.number_format = "#,##0.00" if _row in (4, 6) else "0.00"

    h = ws_sim[f"C{_row}"]
    h.value = _hint
    h.font  = Font(italic=True, size=9, color="888888")
    h.alignment = Alignment(horizontal="left", vertical="center")

ws_sim.row_dimensions[8].height = 8  # spacer

# ── Resultados ───────────────────────────────────────────────────────────────
ws_sim.merge_cells("A9:F9")
ws_sim.row_dimensions[9].height = 24
c = ws_sim["A9"]
c.value = "  RESULTADOS"
c.font  = Font(bold=True, size=11, color="F4EFE2")
c.fill  = PatternFill("solid", fgColor="1B4332")
c.alignment = Alignment(horizontal="left", vertical="center")

_result_rows = [
    (10, "Cuota Mensual",       "=IFERROR(-PMT(B5/100,B7,-B4),0)"),
    (11, "Total a Recibir",     "=B10*B7"),
    (12, "Total de Intereses",  "=B11-B4"),
    (13, "Capital Financiado",  "=B4"),
]
for _row, _label, _formula in _result_rows:
    ws_sim.row_dimensions[_row].height = 26
    c = ws_sim[f"A{_row}"]
    c.value = _label
    c.font  = Font(bold=True, size=11)
    c.fill  = PatternFill("solid", fgColor="EAE7E4")
    c.border = _brd
    c.alignment = Alignment(horizontal="left", vertical="center")
    b = ws_sim[f"B{_row}"]
    b.value = _formula
    b.fill  = PatternFill("solid", fgColor="D4EDDA")
    b.border = _brd
    b.number_format = "#,##0"
    b.font  = Font(bold=True, size=12, color="1B4332")
    b.alignment = Alignment(horizontal="center", vertical="center")

ws_sim.row_dimensions[14].height = 8  # spacer

# ── Tabla de amortización ────────────────────────────────────────────────────
ws_sim.merge_cells("A15:F15")
ws_sim.row_dimensions[15].height = 24
c = ws_sim["A15"]
c.value = "  TABLA DE AMORTIZACIÓN  (se actualiza automáticamente)"
c.font  = Font(bold=True, size=11, color="F4EFE2")
c.fill  = PatternFill("solid", fgColor="3B3028")
c.alignment = Alignment(horizontal="left", vertical="center")

_tbl_heads = ["Cuota #", "Saldo Inicial", "Cuota", "Interés", "Capital", "Saldo Final"]
ws_sim.row_dimensions[16].height = 20
for _ci, _hdr in enumerate(_tbl_heads):
    _col = chr(65 + _ci)
    c = ws_sim[f"{_col}16"]
    c.value = _hdr
    c.font  = Font(bold=True, color="FFFFFF", size=10)
    c.fill  = PatternFill("solid", fgColor="495057")
    c.border = _brd
    c.alignment = Alignment(horizontal="center", vertical="center")

for _n in range(1, 85):
    _r  = 16 + _n
    _bg = "F8F9FA" if _n % 2 == 0 else "FFFFFF"
    _alt = PatternFill("solid", fgColor=_bg)
    ws_sim.row_dimensions[_r].height = 16

    # A: número
    c = ws_sim[f"A{_r}"]
    c.value = _n
    c.fill = _alt; c.border = _brd
    c.font = Font(size=9)
    c.alignment = Alignment(horizontal="center")

    # B: Saldo Inicial
    _bf = "=IF($B$7>=1,$B$4,0)" if _n == 1 else f"=IF(A{_r}<=$B$7,MAX(0,F{_r-1}),0)"
    c = ws_sim[f"B{_r}"]
    c.value = _bf; c.fill = _alt; c.border = _brd
    c.number_format = "#,##0"; c.font = Font(size=9)
    c.alignment = Alignment(horizontal="right")

    # C: Cuota
    c = ws_sim[f"C{_r}"]
    c.value = f"=IF(A{_r}<=$B$7,$B$10,0)"
    c.fill = _alt; c.border = _brd
    c.number_format = "#,##0"; c.font = Font(size=9)
    c.alignment = Alignment(horizontal="right")

    # D: Interés
    c = ws_sim[f"D{_r}"]
    c.value = f"=IF(A{_r}<=$B$7,B{_r}*$B$5/100,0)"
    c.fill = _alt; c.border = _brd
    c.number_format = "#,##0"; c.font = Font(size=9)
    c.alignment = Alignment(horizontal="right")

    # E: Capital amortizado
    c = ws_sim[f"E{_r}"]
    c.value = f"=IF(A{_r}<=$B$7,C{_r}-D{_r},0)"
    c.fill = _alt; c.border = _brd
    c.number_format = "#,##0"; c.font = Font(size=9)
    c.alignment = Alignment(horizontal="right")

    # F: Saldo Final
    c = ws_sim[f"F{_r}"]
    c.value = f"=IF(A{_r}<=$B$7,MAX(0,B{_r}-E{_r}),0)"
    c.fill = _alt; c.border = _brd
    c.number_format = "#,##0"; c.font = Font(size=9)
    c.alignment = Alignment(horizontal="right")

# Nota al pie
_nota_row = 102
ws_sim.merge_cells(f"A{_nota_row}:F{_nota_row}")
c = ws_sim[f"A{_nota_row}"]
c.value = "  Celdas amarillas → entrada manual  |  Celdas verdes → calculadas automáticamente"
c.font  = Font(italic=True, size=9, color="555555")
c.alignment = Alignment(horizontal="left", vertical="center")

# Anchos de columna
ws_sim.column_dimensions["A"].width = 30
ws_sim.column_dimensions["B"].width = 18
ws_sim.column_dimensions["C"].width = 32
ws_sim.column_dimensions["D"].width = 16
ws_sim.column_dimensions["E"].width = 16
ws_sim.column_dimensions["F"].width = 18

# Congelar filas superiores para navegar la tabla sin perder los datos
ws_sim.freeze_panes = "A17"

wb.save(SALIDA)
print(f"Excel creado en: {SALIDA}")
print(f"  Dorlys       : {len(PAGOS_DORLYS)} pagos  → ${len(PAGOS_DORLYS) * 240_000:,}".replace(",","."))
print(f"  Elkin        : {len(PAGOS_ELKIN)} pagos  → ${len(PAGOS_ELKIN) * 240_000:,}".replace(",","."))
print(f"  Juan Andres  : {len(PAGOS_JUAN_ANDRES)} pagos  → ${len(PAGOS_JUAN_ANDRES) * 240_000:,} (1 pendiente)".replace(",","."))
print(f"  Jorge Luis   : {len(PAGOS_JORGE_LUIS)} pagos  → ${len(PAGOS_JORGE_LUIS) * 240_000:,} (FINALIZADO)".replace(",","."))
print(f"  Luis Gabriel : {len(PAGOS_LUIS_GABRIEL)} pagos  → ${len(PAGOS_LUIS_GABRIEL) * 240_000:,} ({72 - len(PAGOS_LUIS_GABRIEL)} pendientes de 72)".replace(",","."))
print(f"  Duvon        : {len(PAGOS_DUVON)} pagos  → ${len(PAGOS_DUVON) * 240_000:,} (1 pendiente de 64)".replace(",","."))
print(f"  Wilmer       : {len(PAGOS_WILMER)} pagos  → ${len(PAGOS_WILMER) * 240_000:,} (5 pendientes de 64)".replace(",","."))
print(f"  Jesus/Yuliana: {len(PAGOS_JESUS)} pagos  → ${len(PAGOS_JESUS) * 240_000:,} (20 pendientes de 72)".replace(",","."))
print(f"  Osmed        : {len(PAGOS_OSMED)} pagos  → ${len(PAGOS_OSMED) * 240_000:,} (10 pendientes de 64)".replace(",","."))
print(f"  Chacal       : {len(PAGOS_CHACAL)} pagos  → ${len(PAGOS_CHACAL) * 240_000:,} (22 pendientes de 72)".replace(",","."))
print(f"  Wilian Jr    : {len(PAGOS_WILIAN_JR)} pagos  → ${len(PAGOS_WILIAN_JR) * 240_000:,} (26 pendientes de 60)".replace(",","."))
print(f"  Darwin       : {len(PAGOS_DARWIN)} pagos  → ${len(PAGOS_DARWIN) * 240_000:,} (10 pendientes de 64)".replace(",","."))
print(f"  Kevin        : {len(PAGOS_KEVIN)} pagos  → ${len(PAGOS_KEVIN) * 240_000:,} (55 pendientes de 84)".replace(",","."))
print(f"  Nabro        : {len(PAGOS_NABRO)} pagos  → ${len(PAGOS_NABRO) * 240_000:,} (13 pendientes de 64)".replace(",","."))
print(f"  Luis Katherin: {len(PAGOS_LUIS_KATHERIN)} pagos  → ${len(PAGOS_LUIS_KATHERIN) * 240_000:,} ({69 - len(PAGOS_LUIS_KATHERIN)} pendientes de 69)".replace(",","."))
print(f"  Lucho        : {len(PAGOS_LUCHO)} pagos  → ${len(PAGOS_LUCHO) * 240_000:,} (13 pendientes de 64)".replace(",","."))
print(f"  Roberto      : {len(PAGOS_ROBERTO)} pagos  → ${len(PAGOS_ROBERTO) * 240_000:,} ({64 - len(PAGOS_ROBERTO)} pendientes de 64)".replace(",","."))
print(f"  Braillon     : {len(PAGOS_BRAILLON)} pagos  → ${len(PAGOS_BRAILLON) * 240_000:,} ({64 - len(PAGOS_BRAILLON)} pendientes de 64)".replace(",","."))
print(f"  Gustavo      : {len(PAGOS_GUSTAVO)} pagos  → ${len(PAGOS_GUSTAVO) * 240_000:,} ({68 - len(PAGOS_GUSTAVO)} pendientes de 68)".replace(",","."))
print(f"  Erzik        : {len(PAGOS_ERZIK)} pagos  → ${len(PAGOS_ERZIK) * TARIFA_ERZIK:,} ({36 - len(PAGOS_ERZIK)} pendientes de 36 | tarifa 384k)".replace(",","."))
print(f"  Efrain       : {len(PAGOS_EFRAIN)} pagos  → ${len(PAGOS_EFRAIN) * 240_000:,} ({84 - len(PAGOS_EFRAIN)} pend de 84 | subtotal real ≈47 cuotas)".replace(",","."))
print(f"  Gonzado      : {len(PAGOS_GONZADO)} pagos  → ${len(PAGOS_GONZADO) * 240_000:,} ({84 - len(PAGOS_GONZADO)} pend de 84 | subtotal cuaderno 48 cuotas)".replace(",","."))
print(f"  Ana Milena   : {len(PAGOS_ANA_MILENA)} pagos  → ${len(PAGOS_ANA_MILENA) * 240_000:,} ({64 - len(PAGOS_ANA_MILENA)} pend de 64)".replace(",","."))
print(f"  Sr Luis      : {len(PAGOS_SR_LUIS)} pagos  → ${len(PAGOS_SR_LUIS) * 240_000:,} ({64 - len(PAGOS_SR_LUIS)} pend de 64)".replace(",","."))
print(f"  Manuel Alga  : {len(PAGOS_MANUEL_ALGA)} pagos  → ${len(PAGOS_MANUEL_ALGA) * 240_000:,} ({72 - len(PAGOS_MANUEL_ALGA)} pend de 72)".replace(",","."))
print(f"  Jorge Estrada: {len(PAGOS_CARLOS_HABITE)} pagos  → ${len(PAGOS_CARLOS_HABITE) * 240_000:,} ({64 - len(PAGOS_CARLOS_HABITE)} pend de 64 | GAP mayo)".replace(",","."))
print(f"  Donita J.C.  : {len(PAGOS_DONITA_JUAN_CAMILO)} pagos  → ${len(PAGOS_DONITA_JUAN_CAMILO) * TARIFA_DONITA:,} ({24 - len(PAGOS_DONITA_JUAN_CAMILO)} pend de 24 | tarifa 420k/mes)".replace(",","."))
print(f"  Alejandra    : {len(PAGOS_ALEJANDRA)} pagos  → ${len(PAGOS_ALEJANDRA) * 240_000:,} ({84 - len(PAGOS_ALEJANDRA)} pend de 84)".replace(",","."))
print(f"  Carlos       : {len(PAGOS_CARLOS)} pagos  → ${len(PAGOS_CARLOS) * 240_000:,} ({72 - len(PAGOS_CARLOS)} pend de 72)".replace(",","."))
print(f"  Francisco    : {len(PAGOS_FRANCISCO)} pagos  → ${len(PAGOS_FRANCISCO) * TARIFA_FRANCISCO:,} ({48 - len(PAGOS_FRANCISCO)} pend de 48 | tarifa 190k bisemanal)".replace(",","."))
print(f"  Yesenia      : {len(PAGOS_YESENIA)} pagos  → ${len(PAGOS_YESENIA) * 240_000:,} ({72 - len(PAGOS_YESENIA)} pend de 72)".replace(",","."))
print(f"  Alvania      : {len(PAGOS_ALVANIA)} pagos  → ${len(PAGOS_ALVANIA) * 240_000:,} ({72 - len(PAGOS_ALVANIA)} pend de 72)".replace(",","."))
print()
print("NOTA: revisa las fechas en el Excel — algunas pueden tener ±1-2 días")
print("      por la letra manuscrita. Edita directamente si algo no cuadra.")
